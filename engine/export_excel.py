"""
Export simulation data to Excel workbooks for ToyLand Distribution.
Creates multiple workbooks in the output folder.
"""
import json
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from collections import defaultdict

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(BASE_DIR)  # parent of engine/
OUTPUT_DIR = os.path.join(PROJECT_DIR, "output", "data")
HIDDEN_DIR = os.path.join(PROJECT_DIR, "output", "hidden_variables")

DATA = None

def _to_str(val):
    """Convert date objects to ISO strings if needed."""
    if hasattr(val, 'isoformat'):
        return val.isoformat()
    return val

def _normalize_dates(obj):
    """Recursively convert date objects to strings in dicts/lists."""
    if isinstance(obj, dict):
        return {k: _normalize_dates(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [_normalize_dates(item) for item in obj]
    elif hasattr(obj, 'isoformat'):
        return obj.isoformat()
    return obj

def load_data(data_dict=None, json_path=None):
    """Load data either from a dict (passed from runner) or from a JSON file."""
    global DATA
    if data_dict:
        DATA = _normalize_dates(data_dict)
    elif json_path:
        with open(json_path) as f:
            DATA = json.load(f)
    else:
        with open(os.path.join(PROJECT_DIR, "sim_output.json")) as f:
            DATA = json.load(f)

# Style constants
HEADER_FONT = Font(name="Arial", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
SUBHEADER_FILL = PatternFill("solid", fgColor="D6E4F0")
SUBHEADER_FONT = Font(name="Arial", bold=True, size=10, color="2F5496")
DATA_FONT = Font(name="Arial", size=10)
MONEY_FMT = '#,##0'
MONEY_FMT_DEC = '#,##0.00'
PCT_FMT = '0.0%'
NUM_FMT = '#,##0'
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)

def style_header_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

def style_data_cell(ws, row, col, fmt=None):
    cell = ws.cell(row=row, column=col)
    cell.font = DATA_FONT
    cell.border = THIN_BORDER
    if fmt:
        cell.number_format = fmt
    return cell

def auto_width(ws, min_w=10, max_w=30):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_w), max_w)

def write_sheet(ws, headers, rows, col_formats=None):
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    for r_idx, row in enumerate(rows, 2):
        for c_idx, val in enumerate(row, 1):
            cell = style_data_cell(ws, r_idx, c_idx)
            cell.value = val
            if col_formats and c_idx - 1 < len(col_formats) and col_formats[c_idx - 1]:
                cell.number_format = col_formats[c_idx - 1]
            if r_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F2F7FB")

    auto_width(ws)

# ============================================================
# 1. COMPANY MASTER DATA WORKBOOK
# ============================================================
def create_master_data():
    wb = Workbook()

    # -- Company Profile Sheet --
    ws = wb.active
    ws.title = "Company Profile"
    company = DATA["company"]
    profile_data = [
        ("Company Name (EN)", company["name"]),
        ("Company Name (TH)", company["name_th"]),
        ("Tax ID", company["tax_id"]),
        ("Founded", company["founded"]),
        ("Industry", company["industry"]),
        ("Headquarters", company["headquarters"]),
        ("Warehouse Address", company["warehouse_address"]),
        ("Employees", company["employees"]),
        ("Registered Capital (THB)", company["registered_capital"]),
        ("Annual Revenue Target (THB)", company["annual_revenue_target"]),
        ("Simulation Start", company["sim_start"]),
        ("Simulation Duration", f"{company['sim_months']} months"),
        ("Currency", company["currency"]),
    ]
    ws.cell(row=1, column=1, value="ToyLand Distribution Co., Ltd.").font = Font(name="Arial", bold=True, size=16, color="2F5496")
    ws.merge_cells("A1:B1")
    ws.cell(row=2, column=1, value="Company Master Data").font = Font(name="Arial", size=12, color="7F7F7F")
    ws.merge_cells("A2:B2")
    for i, (label, val) in enumerate(profile_data, 4):
        ws.cell(row=i, column=1, value=label).font = Font(name="Arial", bold=True, size=10)
        c = ws.cell(row=i, column=2, value=val)
        c.font = DATA_FONT
        if isinstance(val, (int, float)) and "THB" in label:
            c.number_format = MONEY_FMT
    auto_width(ws)

    # -- Products Sheet --
    ws2 = wb.create_sheet("Products")
    cats = DATA["categories"]
    prods = DATA["products"]
    inv = DATA["inventory_params"]
    headers = ["Product ID", "Product Name", "Category", "Category Name", "Unit Cost (THB)", "Selling Price (THB)",
               "Margin %", "Unit", "Weight (kg)", "Reorder Point", "Reorder Qty", "Safety Stock", "Max Stock",
               "Initial Stock WH-01"]
    rows = []
    for p in prods:
        cat_name = cats[p["cat"]]["name"]
        margin = (p["price"] - p["cost"]) / p["price"]
        ip = inv[p["id"]]
        rows.append([p["id"], p["name"], p["cat"], cat_name, p["cost"], p["price"],
                     margin, p["unit"], p["weight_kg"], ip["reorder_point"], ip["reorder_qty"],
                     ip["safety_stock"], ip["max_stock"], ip["initial_wh01"]])
    fmts = [None, None, None, None, MONEY_FMT, MONEY_FMT, PCT_FMT, None, '0.00',
            NUM_FMT, NUM_FMT, NUM_FMT, NUM_FMT, NUM_FMT]
    write_sheet(ws2, headers, rows, fmts)

    # -- Suppliers Sheet --
    ws3 = wb.create_sheet("Suppliers")
    headers = ["Supplier ID", "Supplier Name", "Categories", "Lead Time Min (days)", "Lead Time Max (days)",
               "Reliability", "Min Order (THB)", "Payment Terms (days)"]
    rows = []
    for s in DATA["suppliers"]:
        cat_names = ", ".join([cats[c]["name"] for c in s["categories"]])
        rows.append([s["id"], s["name"], cat_names, s["lead_days"][0], s["lead_days"][1],
                     s["reliability"], s["min_order_thb"], s["payment_terms"]])
    fmts = [None, None, None, None, None, PCT_FMT, MONEY_FMT, None]
    write_sheet(ws3, headers, rows, fmts)

    # -- Customers Sheet --
    ws4 = wb.create_sheet("Customers")
    headers = ["Customer ID", "Customer Name", "Type", "Tier", "Credit Limit (THB)", "Payment Terms (days)",
               "Sales Locations", "Orders/Week", "Primary Categories", "Serve From"]
    rows = []
    for c in DATA["customers"]:
        cat_names = ", ".join([cats[ct]["name"] for ct in c["primary_cats"]])
        locs = c["locations"]
        locs_str = ", ".join(locs) if isinstance(locs, list) else str(locs)
        rows.append([c["id"], c["name"], c["type"], c["tier"], c["credit_limit"], c["payment_days"],
                     locs_str, c["order_freq_per_week"], cat_names, c.get("serve_from", "WH-01")])
    fmts = [None, None, None, None, MONEY_FMT, None, None, '0.0', None, None]
    write_sheet(ws4, headers, rows, fmts)

    # -- Warehouses Sheet --
    ws5 = wb.create_sheet("Warehouses")
    headers = ["Warehouse ID", "Name", "Area (sqm)", "Pallet Positions", "Type", "Products Stored"]
    rows = []
    for w in DATA["warehouses"]:
        prods_str = "All Products" if w["products"] == "all" else ", ".join(w["products"])
        rows.append([w["id"], w["name"], w["sqm"], w["pallets"], w["type"], prods_str])
    write_sheet(ws5, headers, rows)

    # -- Monthly Costs Sheet --
    ws6 = wb.create_sheet("Monthly Costs")
    headers = ["Cost Category", "Monthly Amount (THB)"]
    rows = [[k.replace("_", " ").title(), v] for k, v in DATA["monthly_costs"].items()]
    rows.append(["TOTAL", sum(DATA["monthly_costs"].values())])
    write_sheet(ws6, headers, rows, [None, MONEY_FMT])

    path = os.path.join(OUTPUT_DIR, "01_Master_Data.xlsx")
    wb.save(path)
    print(f"Saved: {path}")

# ============================================================
# 2. ORDERS WORKBOOK
# ============================================================
def create_orders():
    wb = Workbook()
    ws = wb.active
    ws.title = "Order Lines"

    headers = ["Order ID", "Date", "Customer ID", "Product ID", "Qty Ordered", "Qty Filled",
               "Qty Backordered", "Unit Price (THB)", "Discount %", "Line Total (THB)", "COGS (THB)",
               "Warehouse", "Status"]
    rows = []
    for o in DATA["order_log"]:
        rows.append([o["order_id"], o["date"], o["customer_id"], o["product_id"],
                     o["qty_ordered"], o["qty_filled"], o["qty_backordered"],
                     o["unit_price"], o["discount_pct"], o["line_total"], o["cogs"],
                     o.get("fulfill_warehouse", "WH-01"), o["status"]])
    fmts = [None, None, None, None, NUM_FMT, NUM_FMT, NUM_FMT,
            MONEY_FMT_DEC, PCT_FMT, MONEY_FMT_DEC, MONEY_FMT_DEC, None, None]
    write_sheet(ws, headers, rows, fmts)

    # -- Order Summary by Month --
    ws2 = wb.create_sheet("Monthly Summary")
    monthly = defaultdict(lambda: {"revenue": 0, "cogs": 0, "orders": set(), "units": 0, "backorders": 0})
    for o in DATA["order_log"]:
        m = o["date"][:7]
        monthly[m]["revenue"] += o["line_total"]
        monthly[m]["cogs"] += o["cogs"]
        monthly[m]["orders"].add(o["order_id"])
        monthly[m]["units"] += o["qty_filled"]
        monthly[m]["backorders"] += o["qty_backordered"]
    headers = ["Month", "Total Orders", "Units Sold", "Revenue (THB)", "COGS (THB)", "Gross Profit (THB)",
               "Gross Margin %", "Backorder Units"]
    rows = []
    for m in sorted(monthly.keys()):
        d = monthly[m]
        gp = d["revenue"] - d["cogs"]
        gm = gp / d["revenue"] if d["revenue"] > 0 else 0
        rows.append([m, len(d["orders"]), d["units"], round(d["revenue"], 2), round(d["cogs"], 2),
                     round(gp, 2), gm, d["backorders"]])
    fmts = [None, NUM_FMT, NUM_FMT, MONEY_FMT, MONEY_FMT, MONEY_FMT, PCT_FMT, NUM_FMT]
    write_sheet(ws2, headers, rows, fmts)

    # -- Customer Summary --
    ws3 = wb.create_sheet("By Customer")
    cust_data = defaultdict(lambda: {"revenue": 0, "orders": set(), "units": 0})
    for o in DATA["order_log"]:
        cust_data[o["customer_id"]]["revenue"] += o["line_total"]
        cust_data[o["customer_id"]]["orders"].add(o["order_id"])
        cust_data[o["customer_id"]]["units"] += o["qty_filled"]
    headers = ["Customer ID", "Customer Name", "Total Orders", "Units Purchased", "Total Revenue (THB)"]
    cust_map = {c["id"]: c["name"] for c in DATA["customers"]}
    rows = []
    for cid in sorted(cust_data.keys()):
        d = cust_data[cid]
        rows.append([cid, cust_map.get(cid, ""), len(d["orders"]), d["units"], round(d["revenue"], 2)])
    fmts = [None, None, NUM_FMT, NUM_FMT, MONEY_FMT]
    write_sheet(ws3, headers, rows, fmts)

    # -- Product Summary --
    ws4 = wb.create_sheet("By Product")
    prod_data = defaultdict(lambda: {"revenue": 0, "cogs": 0, "units": 0, "backorders": 0})
    for o in DATA["order_log"]:
        prod_data[o["product_id"]]["revenue"] += o["line_total"]
        prod_data[o["product_id"]]["cogs"] += o["cogs"]
        prod_data[o["product_id"]]["units"] += o["qty_filled"]
        prod_data[o["product_id"]]["backorders"] += o["qty_backordered"]
    prod_map = {p["id"]: p["name"] for p in DATA["products"]}
    headers = ["Product ID", "Product Name", "Units Sold", "Revenue (THB)", "COGS (THB)", "Gross Profit (THB)",
               "Margin %", "Units Backordered", "Fill Rate %"]
    rows = []
    for pid in sorted(prod_data.keys()):
        d = prod_data[pid]
        gp = d["revenue"] - d["cogs"]
        gm = gp / d["revenue"] if d["revenue"] > 0 else 0
        total_demand = d["units"] + d["backorders"]
        fill_rate = d["units"] / total_demand if total_demand > 0 else 1.0
        rows.append([pid, prod_map.get(pid, ""), d["units"], round(d["revenue"], 2), round(d["cogs"], 2),
                     round(gp, 2), gm, d["backorders"], fill_rate])
    fmts = [None, None, NUM_FMT, MONEY_FMT, MONEY_FMT, MONEY_FMT, PCT_FMT, NUM_FMT, PCT_FMT]
    write_sheet(ws4, headers, rows, fmts)

    path = os.path.join(OUTPUT_DIR, "02_Orders.xlsx")
    wb.save(path)
    print(f"Saved: {path}")

# ============================================================
# 3. PURCHASE ORDERS WORKBOOK
# ============================================================
def create_purchase_orders():
    wb = Workbook()
    ws = wb.active
    ws.title = "Purchase Orders"

    headers = ["PO ID", "Date", "Supplier ID", "Product ID", "Qty Ordered", "Qty Received",
               "Unit Cost (THB)", "Total Cost (THB)", "Arrival Date", "Lead Days", "Status"]
    rows = []
    for po in DATA["po_log"]:
        rows.append([po["po_id"], po["date"], po["supplier_id"], po["product_id"],
                     po["qty_ordered"], po["qty_received"], po["unit_cost"], po["total_cost"],
                     po["arrival_date"], po["lead_days"], po["status"]])
    fmts = [None, None, None, None, NUM_FMT, NUM_FMT, MONEY_FMT, MONEY_FMT, None, None, None]
    write_sheet(ws, headers, rows, fmts)

    # Supplier performance summary
    ws2 = wb.create_sheet("Supplier Performance")
    sup_data = defaultdict(lambda: {"total_pos": 0, "on_time": 0, "partial": 0, "total_cost": 0, "total_lead": 0})
    for po in DATA["po_log"]:
        sup_data[po["supplier_id"]]["total_pos"] += 1
        sup_data[po["supplier_id"]]["total_cost"] += po["total_cost"]
        sup_data[po["supplier_id"]]["total_lead"] += po["lead_days"]
        if po["status"] == "complete":
            sup_data[po["supplier_id"]]["on_time"] += 1
        elif po["status"] == "partial":
            sup_data[po["supplier_id"]]["partial"] += 1
    sup_map = {s["id"]: s["name"] for s in DATA["suppliers"]}
    headers = ["Supplier ID", "Supplier Name", "Total POs", "Complete Deliveries", "Partial Deliveries",
               "Completion Rate %", "Avg Lead Days", "Total Spend (THB)"]
    rows = []
    for sid in sorted(sup_data.keys()):
        d = sup_data[sid]
        comp_rate = d["on_time"] / d["total_pos"] if d["total_pos"] > 0 else 0
        avg_lead = d["total_lead"] / d["total_pos"] if d["total_pos"] > 0 else 0
        rows.append([sid, sup_map.get(sid, ""), d["total_pos"], d["on_time"], d["partial"],
                     comp_rate, round(avg_lead, 1), d["total_cost"]])
    fmts = [None, None, NUM_FMT, NUM_FMT, NUM_FMT, PCT_FMT, '0.0', MONEY_FMT]
    write_sheet(ws2, headers, rows, fmts)

    path = os.path.join(OUTPUT_DIR, "03_Purchase_Orders.xlsx")
    wb.save(path)
    print(f"Saved: {path}")

# ============================================================
# 4. INVENTORY WORKBOOK
# ============================================================
def create_inventory():
    wb = Workbook()

    # Daily stock is very large, so we'll do end-of-week snapshots for WH-01
    ws = wb.active
    ws.title = "Weekly Stock WH-01"

    # Group by week (every Saturday)
    weekly = {}
    for entry in DATA["daily_stock_log"]:
        if entry.get("warehouse_id", entry.get("location_id", "")) == "WH-01":
            d = entry["date"]
            # Parse date to find Saturdays
            dt = datetime.strptime(d, "%Y-%m-%d")
            if dt.weekday() == 5:  # Saturday
                week_key = d
                if week_key not in weekly:
                    weekly[week_key] = {}
                weekly[week_key][entry["product_id"]] = entry["qty_on_hand"]

    prod_ids = [p["id"] for p in DATA["products"]]
    headers = ["Week Ending"] + [p["id"] for p in DATA["products"]]
    rows = []
    for wk in sorted(weekly.keys()):
        row = [wk] + [weekly[wk].get(pid, 0) for pid in prod_ids]
        rows.append(row)
    fmts = [None] + [NUM_FMT] * len(prod_ids)
    write_sheet(ws, headers, rows, fmts)

    # Transfers
    ws3 = wb.create_sheet("Transfers")
    headers = ["Transfer ID", "Date", "Product ID", "From", "To", "Type", "Qty", "Arrival Date"]
    rows = [[t["transfer_id"], t["date"], t["product_id"],
             t.get("from_loc", t.get("from_wh", "")),
             t.get("to_loc", t.get("to_wh", "")),
             t.get("transfer_type", "WH→WH"),
             t["qty"], t["arrival_date"]]
            for t in DATA["transfer_log"]]
    write_sheet(ws3, headers, rows, [None, None, None, None, None, NUM_FMT, None])

    path = os.path.join(OUTPUT_DIR, "04_Inventory.xlsx")
    wb.save(path)
    print(f"Saved: {path}")

# ============================================================
# 5. FINANCIALS WORKBOOK
# ============================================================
def create_financials():
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily P&L"

    headers = ["Date", "Month", "Revenue (THB)", "COGS (THB)", "Gross Profit (THB)",
               "Fixed Costs (THB)", "Net Profit (THB)", "Orders", "Units Sold", "Stockout Events"]
    rows = []
    for f in DATA["financial_log"]:
        rows.append([f["date"], f["month"], f["revenue"], f["cogs"], f["gross_profit"],
                     f["fixed_costs"], f["net_profit"], f["orders_count"], f["units_sold"], f["stockout_events"]])
    fmts = [None, None, MONEY_FMT_DEC, MONEY_FMT_DEC, MONEY_FMT_DEC,
            MONEY_FMT_DEC, MONEY_FMT_DEC, NUM_FMT, NUM_FMT, NUM_FMT]
    write_sheet(ws, headers, rows, fmts)

    # Monthly P&L
    ws2 = wb.create_sheet("Monthly P&L")
    monthly_fin = defaultdict(lambda: {"revenue": 0, "cogs": 0, "fixed": 0, "orders": 0, "units": 0, "stockouts": 0})
    for f in DATA["financial_log"]:
        m = f["month"]
        monthly_fin[m]["revenue"] += f["revenue"]
        monthly_fin[m]["cogs"] += f["cogs"]
        monthly_fin[m]["fixed"] += f["fixed_costs"]
        monthly_fin[m]["orders"] += f["orders_count"]
        monthly_fin[m]["units"] += f["units_sold"]
        monthly_fin[m]["stockouts"] += f["stockout_events"]

    headers = ["Month", "Revenue (THB)", "COGS (THB)", "Gross Profit (THB)", "Gross Margin %",
               "Fixed Costs (THB)", "Net Profit (THB)", "Net Margin %", "Total Orders", "Units Sold", "Stockout Events"]
    rows = []
    for m in sorted(monthly_fin.keys()):
        d = monthly_fin[m]
        gp = d["revenue"] - d["cogs"]
        gm = gp / d["revenue"] if d["revenue"] > 0 else 0
        np_ = gp - d["fixed"]
        nm = np_ / d["revenue"] if d["revenue"] > 0 else 0
        rows.append([m, round(d["revenue"], 2), round(d["cogs"], 2), round(gp, 2), gm,
                     round(d["fixed"], 2), round(np_, 2), nm, d["orders"], d["units"], d["stockouts"]])
    fmts = [None, MONEY_FMT, MONEY_FMT, MONEY_FMT, PCT_FMT, MONEY_FMT, MONEY_FMT, PCT_FMT, NUM_FMT, NUM_FMT, NUM_FMT]
    write_sheet(ws2, headers, rows, fmts)

    path = os.path.join(OUTPUT_DIR, "05_Financials.xlsx")
    wb.save(path)
    print(f"Saved: {path}")

# ============================================================
# 6. EVENTS LOG WORKBOOK
# ============================================================
def create_events():
    wb = Workbook()
    ws = wb.active
    ws.title = "Events"

    headers = ["Date", "Event Type", "Reference", "Detail"]
    rows = [[e["date"], e["type"], e["ref"], e["detail"]] for e in DATA["event_log"]]
    write_sheet(ws, headers, rows)

    # Summary
    ws2 = wb.create_sheet("Event Summary")
    event_counts = defaultdict(int)
    for e in DATA["event_log"]:
        event_counts[e["type"]] += 1
    headers = ["Event Type", "Count"]
    rows = [[k, v] for k, v in sorted(event_counts.items(), key=lambda x: -x[1])]
    write_sheet(ws2, headers, rows, [None, NUM_FMT])

    path = os.path.join(OUTPUT_DIR, "06_Events_Log.xlsx")
    wb.save(path)
    print(f"Saved: {path}")

# ============================================================
# 7. HIDDEN VARIABLES WORKBOOK (separate folder)
# ============================================================
def create_hidden_variables():
    wb = Workbook()
    ws = wb.active
    ws.title = "Hidden Variables"

    prod_map = {p["id"]: p["name"] for p in DATA["products"]}
    hv = DATA["hidden_vars"]

    headers = ["Product ID", "Product Name", "Base Daily Demand", "Monthly Trend %",
               "Price Sensitivity", "Social Media Buzz", "Competitor Pressure", "Brand Loyalty",
               "Market Sat. Threshold", "Hype Events (month:mult)"]
    rows = []
    for pid in sorted(hv.keys()):
        v = hv[pid]
        hype_str = "; ".join([f"M{h[0]+1}:{h[1]}x" for h in v["hype_events"]]) if v["hype_events"] else "None"
        rows.append([pid, prod_map.get(pid, ""), v["base_daily_demand"], v["trend_monthly_pct"],
                     v["price_sensitivity"], v["social_media_buzz"], v["competitor_pressure"],
                     v["brand_loyalty"], v["market_saturation_threshold"], hype_str])
    fmts = [None, None, NUM_FMT, PCT_FMT, '0.00', '0.00', '0.00', '0.00', NUM_FMT, None]
    write_sheet(ws, headers, rows, fmts)

    # Seasonality matrix
    ws2 = wb.create_sheet("Seasonality Matrix")
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    headers = ["Product ID", "Product Name"] + months
    rows = []
    for pid in sorted(hv.keys()):
        v = hv[pid]
        rows.append([pid, prod_map.get(pid, "")] + v["seasonality_12m"])
    fmts = [None, None] + ['0.00'] * 12
    write_sheet(ws2, headers, rows, fmts)

    # Explanation sheet
    ws3 = wb.create_sheet("Variable Guide")
    explanations = [
        ("base_daily_demand", "Average number of units demanded per working day under normal conditions"),
        ("trend_monthly_pct", "Monthly growth/decline rate. +0.05 = 5% growth per month compounding"),
        ("seasonality_12m", "12 multipliers (Jan-Dec) applied to base demand. 1.0 = normal, 1.5 = 50% boost"),
        ("hype_events", "Specific months where demand spikes (e.g., anime season premiere, new game release)"),
        ("market_saturation_threshold", "Cumulative monthly demand cap before diminishing returns kick in"),
        ("price_sensitivity", "0-1 scale. Higher = customers more price-conscious, discounts have bigger impact"),
        ("social_media_buzz", "0-1 scale. Higher = more viral/trending, amplifies hype events"),
        ("competitor_pressure", "0-1 scale. Higher = more competitor activity, reduces demand by up to 15%"),
        ("brand_loyalty", "0-1 scale. Higher = customers stick to this product despite alternatives"),
    ]
    headers = ["Variable Name", "Description"]
    write_sheet(ws3, headers, explanations)

    path = os.path.join(HIDDEN_DIR, "Hidden_Variables.xlsx")
    wb.save(path)
    print(f"Saved: {path}")

    # Also save as JSON for programmatic use
    json_path = os.path.join(HIDDEN_DIR, "hidden_variables.json")
    with open(json_path, "w") as f:
        json.dump(hv, f, indent=2)
    print(f"Saved: {json_path}")

# ============================================================
# 8. KPI DASHBOARD WORKBOOK
# ============================================================
def create_dashboard():
    wb = Workbook()
    ws = wb.active
    ws.title = "KPI Dashboard"

    # Calculate KPIs
    total_revenue = sum(f["revenue"] for f in DATA["financial_log"])
    total_cogs = sum(f["cogs"] for f in DATA["financial_log"])
    total_fixed = sum(f["fixed_costs"] for f in DATA["financial_log"])
    total_gp = total_revenue - total_cogs
    total_np = total_gp - total_fixed
    total_orders = DATA["summary"]["total_orders"]
    total_units = sum(f["units_sold"] for f in DATA["financial_log"])
    total_stockouts = sum(f["stockout_events"] for f in DATA["financial_log"])

    total_demand = sum(o["qty_ordered"] for o in DATA["order_log"])
    total_filled = sum(o["qty_filled"] for o in DATA["order_log"])
    fill_rate = total_filled / total_demand if total_demand > 0 else 0

    orders_with_backorder = len(set(o["order_id"] for o in DATA["order_log"] if o["qty_backordered"] > 0))
    orders_total = len(set(o["order_id"] for o in DATA["order_log"]))
    order_completion = 1 - (orders_with_backorder / orders_total) if orders_total > 0 else 1

    # Title
    ws.cell(row=1, column=1, value="ToyLand Distribution — 12-Month KPI Dashboard").font = Font(name="Arial", bold=True, size=16, color="2F5496")
    ws.merge_cells("A1:D1")
    ws.cell(row=2, column=1, value=f"Period: {DATA['company']['sim_start']} to 12 months").font = Font(name="Arial", size=11, color="7F7F7F")

    # Financial KPIs
    kpis = [
        ("", ""),
        ("FINANCIAL PERFORMANCE", ""),
        ("Total Revenue", total_revenue, MONEY_FMT),
        ("Total COGS", total_cogs, MONEY_FMT),
        ("Gross Profit", total_gp, MONEY_FMT),
        ("Gross Margin", total_gp / total_revenue if total_revenue else 0, PCT_FMT),
        ("Total Fixed Costs", total_fixed, MONEY_FMT),
        ("Net Profit", total_np, MONEY_FMT),
        ("Net Margin", total_np / total_revenue if total_revenue else 0, PCT_FMT),
        ("", ""),
        ("OPERATIONS", ""),
        ("Total Orders", total_orders, NUM_FMT),
        ("Total Units Sold", total_units, NUM_FMT),
        ("Avg Order Value (THB)", total_revenue / total_orders if total_orders else 0, MONEY_FMT),
        ("Avg Units per Order", total_units / total_orders if total_orders else 0, '0.0'),
        ("", ""),
        ("SERVICE LEVEL", ""),
        ("Unit Fill Rate", fill_rate, PCT_FMT),
        ("Order Completion Rate", order_completion, PCT_FMT),
        ("Total Stockout Events", total_stockouts, NUM_FMT),
        ("", ""),
        ("PROCUREMENT", ""),
        ("Total Purchase Orders", DATA["summary"]["total_pos"], NUM_FMT),
        ("Total Transfers (WH→WH)", DATA["summary"]["total_transfers"], NUM_FMT),
        ("Working Days Simulated", DATA["summary"]["total_days"], NUM_FMT),
    ]

    row = 4
    for item in kpis:
        if len(item) == 2 and item[1] == "":
            if item[0]:
                ws.cell(row=row, column=1, value=item[0]).font = Font(name="Arial", bold=True, size=12, color="2F5496")
                ws.cell(row=row, column=1).fill = SUBHEADER_FILL
                ws.cell(row=row, column=2).fill = SUBHEADER_FILL
            row += 1
            continue
        label, val = item[0], item[1]
        fmt = item[2] if len(item) > 2 else None
        ws.cell(row=row, column=1, value=label).font = Font(name="Arial", size=10)
        c = ws.cell(row=row, column=2, value=val)
        c.font = Font(name="Arial", bold=True, size=11)
        if fmt:
            c.number_format = fmt
        row += 1

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 18

    path = os.path.join(OUTPUT_DIR, "07_KPI_Dashboard.xlsx")
    wb.save(path)
    print(f"Saved: {path}")

# ============================================================
# RUN ALL EXPORTS
# ============================================================
def export_all(data_dict=None, json_path=None):
    """Main export function. Call with data_dict from runner, or json_path, or standalone."""
    load_data(data_dict, json_path)
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    os.makedirs(HIDDEN_DIR, exist_ok=True)
    create_master_data()
    create_orders()
    create_purchase_orders()
    create_inventory()
    create_financials()
    create_events()
    create_hidden_variables()
    create_dashboard()
    print("\nAll workbooks exported successfully!")

if __name__ == "__main__":
    export_all()
