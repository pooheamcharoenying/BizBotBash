"""
ToyLand Distribution Co., Ltd. - Config-Driven Simulation Engine v2
====================================================================
Two operating modes:
  AUTO  – fully autonomous; all POs, transfers, refills follow simple rules
  BOT   – step-by-step; an external controller injects commands each day

Price elasticity model:
  demand_multiplier = (1 - discount_pct) ^ (-elasticity)
  elasticity = price_sensitivity × 3  (range ≈ 0.45–2.1)
  Revenue per unit = price × (1 - discount_pct)
  Default: discount_pct = 0 everywhere (no discounts)

Excel output:
  initial_state.xlsx  – produced once at sim start
  month_YYYY-MM.xlsx  – produced at end of each calendar month

Bot commands (list of dicts passed to step_day):
  {"action": "issue_po",   "supplier_id", "product_id", "qty"}
  {"action": "transfer",   "from_loc", "to_loc", "product_id", "qty"}
  {"action": "set_discount","product_id", "discount_pct", "location_id"?}
  {"action": "set_shelf",  "location_id", "product_id", "shelf_grade"}
"""
import random
import math
import json
import os
from datetime import date, timedelta
from collections import defaultdict

from datetime import datetime as _datetime


def _popularity_cycle(pid, mi, hv):
    """Cyclical popularity multiplier for a product at month index mi.

    Each product oscillates around its baseline on an 18-36 month cycle
    with a deterministic phase derived from its ID, so "hot" and "cold"
    products rotate naturally over time instead of monotonically decaying.
    """
    rng = random.Random(pid)
    period = rng.randint(18, 36)
    phase = rng.uniform(0, 2 * math.pi)
    amplitude = min(0.3 + abs(hv.get("trend_monthly_pct", 0)) * 5, 0.55)
    return 1 + amplitude * math.sin(2 * math.pi * mi / period + phase)


# ─────────────────────────────────────────────────────────────
# Random trend-event configuration (hidden from bots)
# ─────────────────────────────────────────────────────────────
# Trends can fire either globally (affects the target everywhere) or
# locally (affects the target only at one specific store).
# Local rates are much smaller because they roll per (target × location)
# combination, so the total event volume stays balanced.
#
# For a given (product, location) pair the effective multiplier is:
#   1 + Σ (direction × magnitude) over every active event that matches
# (additive score combining, per user spec) — capped at ×0.25..×4.0.
TREND_SCOPES = [
    # (scope_name, per_location, rate)
    ("location",         False, 0.15),  # inherently local (the location IS the target)
    ("product_global",   False, 0.04),  # per SKU
    ("product_local",    True,  0.005), # per (SKU × location)
    ("brand_global",     False, 0.08),  # per supplier (== brand for now)
    ("brand_local",      True,  0.02),  # per (supplier × location)
    ("category_global",  False, 0.10),  # per category
    ("category_local",   True,  0.03),  # per (category × location)
]
# (probability, min duration months, max duration months)
TREND_DURATION_BUCKETS = [
    (0.40, 1, 3),    # short-term fad
    (0.40, 4, 6),    # mid-term
    (0.20, 7, 18),   # long-term
]
TREND_MAG_RANGE = (0.10, 0.50)   # ±10% to ±50%
TREND_MULT_CAP = (0.25, 4.0)     # soft cap so compounded trends don't explode
TREND_PERMANENT_PROB = 0.20      # 20% of trends persist forever, 80% revert
TREND_PERMANENT_SENTINEL = 99999 # end_month for permanent events (never prunes)

# Area per shelf slot (cm²). Combined with slots_per_shelf this gives
# each shelf a total area, which we divide by a product's base_area_cm2
# to compute how many units of that product fit on one shelf.
SLOT_AREA_CM2 = 200

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(BASE_DIR)
CONFIG_DIR = os.path.join(PROJECT_DIR, "config")
DATA_DIR = os.path.join(PROJECT_DIR, "data")
EXCEL_DIR = os.path.join(PROJECT_DIR, "sim_excel")  # legacy default


def load_json(filename):
    with open(os.path.join(CONFIG_DIR, filename)) as f:
        return json.load(f)


def load_config():
    cfg = {}
    cfg["company"] = load_json("company.json")
    cfg["categories"] = load_json("categories.json")
    cfg["products"] = load_json("products.json")
    cfg["suppliers"] = load_json("suppliers.json")
    cfg["customers"] = []
    cfg["warehouses"] = load_json("warehouses.json")
    cfg["sales_locations"] = load_json("sales_locations.json")
    cfg["costs"] = load_json("costs.json")
    cfg["shelf_config"] = load_json("shelf_config.json")

    raw_hv = load_json("hidden_variables.json")
    templates = raw_hv.pop("_seasonality_templates", {})
    raw_hv.pop("_variable_guide", None)
    raw_hv.pop("_composite_score_guide", None)

    hidden_vars = {}
    for pid, hv in raw_hv.items():
        season = hv["seasonality"]
        if isinstance(season, str):
            hv["seasonality_12m"] = templates.get(season, [1.0]*12)
        else:
            hv["seasonality_12m"] = season
        hidden_vars[pid] = hv
    cfg["hidden_vars"] = hidden_vars

    # Product→supplier mapping: pick supplier with shortest avg lead time per category
    # (matches demo_baseline_bot logic)
    cat_suppliers = defaultdict(list)
    for s in cfg["suppliers"]:
        for cat in s["categories"]:
            cat_suppliers[cat].append(s)

    product_supplier = {}
    for p in cfg["products"]:
        cat = p["cat"]
        candidates = cat_suppliers.get(cat, [])
        if candidates:
            best = min(candidates, key=lambda s: sum(s["lead_days"]) / 2)
            product_supplier[p["id"]] = best["id"]
        else:
            product_supplier[p["id"]] = cfg["suppliers"][-1]["id"]
    cfg["product_supplier"] = product_supplier

    physical_locs = [l for l in cfg["sales_locations"] if l["type"] != "Online"]
    online_locs = [l for l in cfg["sales_locations"] if l["type"] == "Online"]
    cfg["physical_locs"] = physical_locs
    cfg["online_locs"] = online_locs

    region_map = defaultdict(list)
    for l in physical_locs:
        region_map[l["region"]].append(l["id"])
    cfg["region_map"] = dict(region_map)

    sc = cfg["shelf_config"]
    cfg["shelf_mult"] = sc["shelf_multiplier"]
    cfg["location_mult"] = sc["location_multiplier"]
    cfg["product_mult"] = sc["product_multiplier"]
    cfg["location_grade"] = {l["id"]: l.get("location_grade", "B") for l in physical_locs}
    cfg["product_grade"] = {pid: hv.get("product_grade", "B") for pid, hv in hidden_vars.items()}

    cat_priority = sc["category_shelf_priority"]
    a_cats = set(cat_priority.get("A_shelf_categories", []))
    b_cats = set(cat_priority.get("B_shelf_categories", []))

    # Initial shelf assignment uses physical ABC shelves.
    # Bots can later assign any grade (A-E) via set_shelf command.
    product_shelf_grade = {}
    for loc in physical_locs:
        loc_id = loc["id"]
        shelf_grades = loc.get("shelf_grades", ["B"])
        sps = loc.get("slots_per_shelf", 30)
        a_capacity = shelf_grades.count("A") * sps
        b_capacity = shelf_grades.count("B") * sps

        a_prefs = [p for p in cfg["products"] if p["cat"] in a_cats]
        b_prefs = [p for p in cfg["products"] if p["cat"] in b_cats]
        c_prefs = [p for p in cfg["products"] if p["cat"] not in a_cats and p["cat"] not in b_cats]

        a_used = b_used = 0
        for p in a_prefs:
            if a_used < a_capacity:
                product_shelf_grade[(loc_id, p["id"])] = "A"; a_used += 1
            elif b_used < b_capacity:
                product_shelf_grade[(loc_id, p["id"])] = "B"; b_used += 1
            else:
                product_shelf_grade[(loc_id, p["id"])] = "C"
        for p in b_prefs:
            if b_used < b_capacity:
                product_shelf_grade[(loc_id, p["id"])] = "B"; b_used += 1
            elif a_used < a_capacity:
                product_shelf_grade[(loc_id, p["id"])] = "A"; a_used += 1
            else:
                product_shelf_grade[(loc_id, p["id"])] = "C"
        for p in c_prefs:
            product_shelf_grade[(loc_id, p["id"])] = "C"

    cfg["product_shelf_grade"] = product_shelf_grade

    supplier_map = {s["id"]: s for s in cfg["suppliers"]}
    inventory_params = {}
    rng = random.Random(cfg["company"].get("random_seed", 2026))

    num_stores = len(physical_locs)
    for p in cfg["products"]:
        # Default bot uses ONLY public information. Order-up-to target
        # is one full refill round (refill_num × num_stores). No fixed
        # safety cushion — bots that layer demand-aware buffers on top
        # will visibly beat this.
        refill = p.get("refill_num", 5)
        # Target WH = 1.75× one refill round — a measured buffer that
        # covers supplier lead time (5-10 days) and the ~2% yield loss
        # to partial deliveries, without the runaway 2× accumulation
        # the old fixed rule caused under the new shelf mechanic.
        base_round = refill * num_stores
        target_wh = int(base_round * 1.75)
        trigger_wh = max(refill, target_wh // 2)
        initial_wh01 = target_wh

        inventory_params[p["id"]] = {
            "target_wh": max(target_wh, 10),
            "trigger_wh": max(trigger_wh, 2),
            # legacy field: keep for any caller that still reads reorder_qty;
            # treat it as the target amount (order-up-to style)
            "reorder_qty": max(target_wh, 10),
            "initial_wh01": max(initial_wh01, 10),
        }
    cfg["inventory_params"] = inventory_params
    cfg["discount_tiers"] = cfg["costs"]["discount_tiers"]
    cfg["monthly_costs"] = cfg["costs"]["monthly_fixed_costs"]
    cfg["variable_costs"] = cfg["costs"].get("variable_costs", {})
    return cfg


# ═══════════════════════════════════════════════════════════
# EXCEL OUTPUT
# ═══════════════════════════════════════════════════════════
def _ensure_dir(d=None):
    os.makedirs(d or EXCEL_DIR, exist_ok=True)


def _style_header(ws, cols):
    from openpyxl.styles import Font, PatternFill, Alignment
    hdr_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="2F5496")
    hdr_align = Alignment(horizontal="center", vertical="center")
    for c in range(1, cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align


def write_initial_state_excel(engine, out_dir=None):
    """Produce initial_state.xlsx with sheets: Products, Locations, Suppliers, WH_Stock, Store_Stock, Discounts."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, numbers

    target = out_dir or EXCEL_DIR
    _ensure_dir(target)
    wb = Workbook()

    # --- Products sheet ---
    ws = wb.active
    ws.title = "Products"
    headers = ["Product ID", "Name", "Category", "Cost (THB)", "Price (THB)", "Grade",
               "Base Demand", "Refill Num", "Volume cm3", "Supplier"]
    ws.append(headers)
    _style_header(ws, len(headers))
    for p in engine.cfg["products"]:
        hv = engine.cfg["hidden_vars"].get(p["id"], {})
        ws.append([p["id"], p["name"], p["cat"], p["cost"], p["price"],
                   hv.get("product_grade", "B"), hv.get("base_daily_demand", 0),
                   p.get("refill_num", 5), p.get("volume_cm3", 5000),
                   engine.cfg["product_supplier"].get(p["id"], "")])
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 10

    # --- Locations sheet ---
    ws2 = wb.create_sheet("Locations")
    headers = ["Location ID", "Name", "Type", "Region", "Grade", "Traffic", "Conv Rate",
               "Shelves", "Slots/Shelf", "Capacity m3"]
    ws2.append(headers)
    _style_header(ws2, len(headers))
    for loc in engine.cfg["sales_locations"]:
        ws2.append([loc["id"], loc["name"], loc["type"], loc.get("region", ""),
                    loc.get("location_grade", ""), loc.get("daily_foot_traffic", 0),
                    loc.get("conversion_rate", 0), loc.get("shelves", 0),
                    loc.get("slots_per_shelf", 0), loc.get("storage_capacity_m3", 0)])
    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 30

    # --- Suppliers sheet ---
    ws3 = wb.create_sheet("Suppliers")
    headers = ["Supplier ID", "Name", "Categories", "Lead Min", "Lead Max",
               "Reliability", "Min Order THB", "Payment Terms"]
    ws3.append(headers)
    _style_header(ws3, len(headers))
    for s in engine.cfg["suppliers"]:
        ws3.append([s["id"], s["name"], ", ".join(s["categories"]),
                    s["lead_days"][0], s["lead_days"][1], s["reliability"],
                    s["min_order_thb"], s["payment_terms"]])
    ws3.column_dimensions["B"].width = 30

    # --- WH Stock sheet ---
    ws4 = wb.create_sheet("WH_Stock")
    headers = ["Product ID", "Product Name", "Qty on Hand"]
    ws4.append(headers)
    _style_header(ws4, len(headers))
    for p in engine.cfg["products"]:
        qty = engine.stock.get(("WH-01", p["id"]), 0)
        ws4.append([p["id"], p["name"], qty])
    ws4.column_dimensions["B"].width = 40

    # --- Store Stock sheet ---
    ws5 = wb.create_sheet("Store_Stock")
    store_names = {loc["id"]: loc["name"] for loc in engine.cfg["physical_locs"]}
    store_ids = [loc["id"] for loc in engine.cfg["physical_locs"]]
    headers = ["Product ID", "Product Name"] + [store_names[sid] for sid in store_ids]
    ws5.append(headers)
    _style_header(ws5, len(headers))
    for p in engine.cfg["products"]:
        row = [p["id"], p["name"]]
        for sid in store_ids:
            row.append(engine.stock.get((sid, p["id"]), 0))
        ws5.append(row)
    ws5.column_dimensions["B"].width = 40
    for i in range(3, 3 + len(store_ids)):
        ws5.column_dimensions[ws5.cell(row=1, column=i).column_letter].width = 18

    # --- Discounts sheet ---
    ws6 = wb.create_sheet("Discounts")
    headers = ["Product ID", "Product Name", "Location ID", "Discount %", "Effective Price"]
    ws6.append(headers)
    _style_header(ws6, len(headers))
    for key, disc in engine.discounts.items():
        pid, loc_id = key
        p = engine.product_map[pid]
        eff_price = round(p["price"] * (1 - disc), 2)
        ws6.append([pid, p["name"], loc_id or "ALL", f"{disc*100:.1f}%", eff_price])
    ws6.column_dimensions["B"].width = 40

    path = os.path.join(target, "initial_state.xlsx")
    wb.save(path)
    print(f"  Excel: {path}")
    return path


def write_monthly_excel(engine, month_str, out_dir=None):
    """Produce month_YYYY-MM.xlsx with sheets: Sales, Stock_Snapshot, POs, Transfers, Financials, Discounts, Action_Log."""
    from openpyxl import Workbook
    target = out_dir or EXCEL_DIR
    _ensure_dir(target)
    wb = Workbook()

    # --- Sales sheet ---
    ws = wb.active
    ws.title = "Sales"
    headers = ["Date", "Order ID", "Customer", "Product ID", "Product Name", "Location",
               "Qty Ordered", "Qty Filled", "Unit Price", "Discount %", "Line Total", "COGS", "Status", "Source"]
    ws.append(headers)
    _style_header(ws, len(headers))
    month_orders = [o for o in engine.order_log if str(o["date"])[:7] == month_str]
    for o in month_orders:
        pname = engine.product_map.get(o["product_id"], {}).get("name", "")
        ws.append([str(o["date"]), o["order_id"], o["customer_id"], o["product_id"], pname,
                   o["sales_location_name"], o["qty_ordered"], o["qty_filled"],
                   o["unit_price"], f"{o['discount_pct']*100:.1f}%", o["line_total"],
                   o["cogs"], o["status"], o.get("source", "auto")])
    ws.column_dimensions["E"].width = 35

    # --- Stock Snapshot (end of month) ---
    ws2 = wb.create_sheet("Stock_Snapshot")
    all_locs = ["WH-01"] + [loc["id"] for loc in engine.cfg["physical_locs"]]
    loc_names = {"WH-01": "Warehouse"}
    loc_names.update({loc["id"]: loc["name"] for loc in engine.cfg["physical_locs"]})
    headers = ["Product ID", "Product Name"] + [loc_names[lid] for lid in all_locs]
    ws2.append(headers)
    _style_header(ws2, len(headers))
    for p in engine.cfg["products"]:
        row = [p["id"], p["name"]]
        for lid in all_locs:
            row.append(engine.stock.get((lid, p["id"]), 0))
        ws2.append(row)
    ws2.column_dimensions["B"].width = 35

    # --- POs sheet ---
    ws3 = wb.create_sheet("Purchase_Orders")
    headers = ["Date", "PO ID", "Supplier", "Product ID", "Product Name",
               "Qty Ordered", "Qty Received", "Unit Cost", "Total Cost",
               "Lead Days", "Status", "Source"]
    ws3.append(headers)
    _style_header(ws3, len(headers))
    month_pos = [po for po in engine.po_log if str(po["date"])[:7] == month_str]
    for po in month_pos:
        pname = engine.product_map.get(po["product_id"], {}).get("name", "")
        ws3.append([str(po["date"]), po["po_id"], po["supplier_id"], po["product_id"],
                    pname, po["qty_ordered"], po["qty_received"], po["unit_cost"],
                    po["total_cost"], po["lead_days"], po["status"],
                    po.get("source", "auto")])
    ws3.column_dimensions["E"].width = 35

    # --- Transfers sheet ---
    ws4 = wb.create_sheet("Transfers")
    headers = ["Date", "Transfer ID", "Product ID", "Product Name", "From", "To",
               "Type", "Qty", "Source"]
    ws4.append(headers)
    _style_header(ws4, len(headers))
    month_tr = [t for t in engine.transfer_log if str(t["date"])[:7] == month_str]
    for t in month_tr:
        pname = engine.product_map.get(t["product_id"], {}).get("name", "")
        ws4.append([str(t["date"]), t["transfer_id"], t["product_id"], pname,
                    t.get("from_loc", ""), t.get("to_loc", ""), t.get("transfer_type", ""),
                    t["qty"], t.get("source", "auto")])
    ws4.column_dimensions["D"].width = 35

    # --- Financials sheet ---
    ws5 = wb.create_sheet("Financials")
    headers = ["Date", "Revenue", "COGS", "Gross Profit", "Fixed Costs", "Net Profit",
               "Orders", "Units Sold", "Stockouts"]
    ws5.append(headers)
    _style_header(ws5, len(headers))
    month_fin = [f for f in engine.financial_log if f["month"] == month_str]
    for f in month_fin:
        ws5.append([str(f["date"]), f["revenue"], f["cogs"], f["gross_profit"],
                    f["fixed_costs"], f["net_profit"], f["orders_count"],
                    f["units_sold"], f["stockout_events"]])

    # --- Active Discounts ---
    ws6 = wb.create_sheet("Active_Discounts")
    headers = ["Product ID", "Product Name", "Location", "Discount %", "Effective Price"]
    ws6.append(headers)
    _style_header(ws6, len(headers))
    for key, disc in engine.discounts.items():
        if disc > 0:
            pid, loc_id = key
            p = engine.product_map.get(pid, {})
            eff_price = round(p.get("price", 0) * (1 - disc), 2)
            ws6.append([pid, p.get("name", ""), loc_id or "ALL",
                        f"{disc*100:.1f}%", eff_price])
    ws6.column_dimensions["B"].width = 35

    # --- Action Log ---
    ws7 = wb.create_sheet("Action_Log")
    headers = ["Date", "Source", "Action", "Details"]
    ws7.append(headers)
    _style_header(ws7, len(headers))
    month_actions = [a for a in engine.action_log if str(a["date"])[:7] == month_str]
    for a in month_actions:
        ws7.append([str(a["date"]), a["source"], a["action"], a["details"]])
    ws7.column_dimensions["D"].width = 60

    path = os.path.join(target, f"month_{month_str}.xlsx")
    wb.save(path)
    print(f"  Excel: {path}")
    return path


def save_run(engine, label="baseline", compact_data=None):
    """Save a simulation run to data/{label}_{timestamp}/.
    Writes Excel files + run_meta.json. Returns the run folder path."""
    ts = _datetime.now().strftime("%Y-%m-%d_%H%M%S")
    safe_label = label.replace(" ", "_").replace("/", "-")
    folder_name = f"{safe_label}_{ts}"
    run_dir = os.path.join(DATA_DIR, folder_name)
    os.makedirs(run_dir, exist_ok=True)

    # Write Excel files into this run folder
    write_initial_state_excel(engine, out_dir=run_dir)
    # Write all monthly Excel files
    months_written = set()
    for entry in engine.financial_log:
        m = entry["month"]
        if m not in months_written:
            write_monthly_excel(engine, m, out_dir=run_dir)
            months_written.add(m)

    # Write run metadata
    meta = {
        "label": label,
        "folder": folder_name,
        "timestamp": ts,
        "mode": engine.mode,
        "seed": engine.cfg["company"].get("random_seed"),
        "months": engine.cfg["company"].get("sim_months", 12),
        "total_days": engine.day_count,
        "total_revenue": round(engine.total_revenue, 2),
        "total_cogs": round(engine.total_cogs, 2),
        "gross_profit": round(engine.total_revenue - engine.total_cogs, 2),
        "total_orders": engine.order_counter,
        "total_pos": engine.po_counter,
        "total_transfers": engine.transfer_counter,
    }

    # Save compact dashboard JSON if provided
    if compact_data:
        with open(os.path.join(run_dir, "compact.json"), "w") as f:
            json.dump(compact_data, f, separators=(',', ':'), default=str)
        meta["has_compact"] = True

    with open(os.path.join(run_dir, "run_meta.json"), "w") as f:
        json.dump(meta, f, indent=2, default=str)

    print(f"  Run saved: {run_dir}")
    return run_dir


def list_runs():
    """List all saved runs from data/ folder, newest first."""
    if not os.path.isdir(DATA_DIR):
        return []
    runs = []
    for name in sorted(os.listdir(DATA_DIR), reverse=True):
        meta_path = os.path.join(DATA_DIR, name, "run_meta.json")
        if os.path.isfile(meta_path):
            with open(meta_path) as f:
                meta = json.load(f)
            runs.append(meta)
    return runs


def build_compact(engine):
    """Build the compact data structure the dashboard expects."""
    D = engine.get_output()
    compact = {}

    if isinstance(D["categories"], dict):
        compact["categories"] = {cid: cv["name"] for cid, cv in D["categories"].items()}
    else:
        compact["categories"] = {c["id"]: c["name"] for c in D["categories"]}

    hv = D.get("hidden_vars", {})

    compact["products"] = [{"id": p["id"], "name": p["name"], "cat": p["cat"],
                             "cost": p["cost"], "price": p["price"],
                             "grade": hv.get(p["id"], {}).get("product_grade", "B"),
                             "score": hv.get(p["id"], {}).get("composite_score", 0),
                             "demand": hv.get(p["id"], {}).get("base_daily_demand", 0),
                             "buzz": hv.get(p["id"], {}).get("social_media_buzz", 0),
                             "loyalty": hv.get(p["id"], {}).get("brand_loyalty", 0),
                             "trend": hv.get(p["id"], {}).get("trend_monthly_pct", 0)}
                            for p in D["products"]]

    compact["locations"] = [{"id": l["id"], "name": l["name"], "type": l["type"],
                              "region": l["region"],
                              "grade": l.get("location_grade", ""),
                              "shelves": l.get("shelves", 0),
                              "shelf_grades": l.get("shelf_grades", []),
                              "traffic": l.get("daily_foot_traffic", 0),
                              "conv": l.get("conversion_rate", 0),
                              "address": l.get("address", ""),
                              "hours": l.get("operating_hours", ""),
                              "capacity_m3": l.get("storage_capacity_m3", 0)}
                             for l in D["sales_locations"]]

    compact["suppliers"] = [{"id": s["id"], "name": s["name"],
                              "categories": s.get("categories", []),
                              "lead_days": s.get("lead_days", [5, 10]),
                              "reliability": s.get("reliability", 0.9),
                              "min_order_thb": s.get("min_order_thb", 20000),
                              "payment_terms": s.get("payment_terms", 30)}
                             for s in D["suppliers"]]

    sales_agg = defaultdict(lambda: {"q": 0, "r": 0})
    for o in D["order_log"]:
        m = str(o["date"])[:7]
        key = (m, o["product_id"], o["sales_location_id"])
        sales_agg[key]["q"] += o["qty_filled"]
        sales_agg[key]["r"] += o["line_total"]
    compact["sales"] = [{"m": k[0], "p": k[1], "l": k[2],
                          "q": v["q"], "r": round(v["r"], 2)}
                         for k, v in sales_agg.items() if v["q"] > 0]

    orders_by_ml = defaultdict(set)
    for o in D["order_log"]:
        if o["qty_filled"] > 0:
            m = str(o["date"])[:7]
            orders_by_ml[(m, o["sales_location_id"])].add(o["order_id"])
    compact["orders_by_ml"] = [{"m": k[0], "l": k[1], "c": len(v)}
                                for k, v in orders_by_ml.items()]

    stock_snap = {}
    for entry in D["daily_stock_log"]:
        loc_id = entry.get("location_id", entry.get("warehouse_id", ""))
        if loc_id == "WH-01":
            m = str(entry["date"])[:7]
            pid = entry["product_id"]
            stock_snap[(m, pid)] = entry["qty_on_hand"]
    compact["stock"] = [{"m": k[0], "p": k[1], "s": v} for k, v in stock_snap.items()]

    po_agg = defaultdict(lambda: {"count": 0, "qty": 0, "cost": 0})
    for po in D["po_log"]:
        m = str(po["date"])[:7]
        po_agg[m]["count"] += 1
        po_agg[m]["qty"] += po["qty_ordered"]
        po_agg[m]["cost"] += po["total_cost"]
    compact["po_agg"] = [{"m": m, "count": v["count"],
                           "qty": v["qty"], "spend": round(v["cost"], 2)}
                          for m, v in sorted(po_agg.items())]

    compact["po_detail"] = [{"id": po["po_id"], "d": str(po["date"]),
                              "s": po["supplier_id"], "p": po["product_id"],
                              "qo": po["qty_ordered"],
                              "qr": po.get("qty_received", 0),
                              "cost": round(po["total_cost"], 2),
                              "lead": po.get("lead_days", 0),
                              "st": po.get("status", "pending")}
                             for po in D["po_log"]]

    tr_agg = defaultdict(lambda: {"count": 0, "qty": 0, "shipments": set()})
    for t in D["transfer_log"]:
        m = str(t["date"])[:7]
        tr_agg[m]["count"] += 1
        tr_agg[m]["qty"] += t["qty"]
        tr_agg[m]["shipments"].add(t["transfer_id"])
    compact["tr_agg"] = [{"m": m, "count": v["count"], "qty": v["qty"],
                           "shipments": len(v["shipments"])}
                          for m, v in sorted(tr_agg.items())]

    sorted_tr = sorted(D["transfer_log"], key=lambda t: str(t["date"]), reverse=True)[:500]
    compact["tr_detail"] = [{"id": t["transfer_id"], "d": str(t["date"]),
                              "p": t["product_id"],
                              "from": t.get("from_loc", t.get("from_wh", "")),
                              "to": t.get("to_loc", t.get("to_wh", "")),
                              "type": t.get("transfer_type", ""),
                              "qty": t["qty"]}
                             for t in sorted_tr]

    total_units = sum(s["q"] for s in compact["sales"])
    total_rev = sum(s["r"] for s in compact["sales"])

    # ── Monthly financials (P&L) ──
    monthly_fin = defaultdict(lambda: {"rev": 0, "cogs": 0, "gp": 0, "fixed": 0, "var": 0, "net": 0, "units": 0, "orders": 0, "stockouts": 0})
    for f in D.get("financial_log", []):
        m = f["month"]
        monthly_fin[m]["rev"] += f["revenue"]
        monthly_fin[m]["cogs"] += f["cogs"]
        monthly_fin[m]["gp"] += f["gross_profit"]
        monthly_fin[m]["fixed"] += f["fixed_costs"]
        monthly_fin[m]["var"] += f.get("variable_costs", 0)
        monthly_fin[m]["net"] += f["net_profit"]
        monthly_fin[m]["units"] += f["units_sold"]
        monthly_fin[m]["orders"] += f["orders_count"]
        monthly_fin[m]["stockouts"] += f.get("stockout_events", 0)
    compact["financials"] = [{"m": m, "rev": round(v["rev"]), "cogs": round(v["cogs"]),
                               "gp": round(v["gp"]), "fixed": round(v["fixed"]),
                               "var": round(v["var"]),
                               "net": round(v["net"]), "units": v["units"],
                               "orders": v["orders"], "stockouts": v["stockouts"]}
                              for m, v in sorted(monthly_fin.items())]

    # ── Monthly cost breakdown (fixed + variable costs config) ──
    compact["monthly_costs"] = D.get("monthly_costs", engine.cfg.get("monthly_costs", {}))
    compact["variable_costs_config"] = engine.cfg.get("variable_costs", {})

    total_fixed = sum(f["fixed"] for f in compact["financials"])
    total_var = sum(f["var"] for f in compact["financials"])
    total_cogs = sum(f["cogs"] for f in compact["financials"])
    total_net = sum(f["net"] for f in compact["financials"])

    # ── Ending inventory value (at cost) ──
    product_cost = {p["id"]: p["cost"] for p in D["products"]}
    ending_inventory_value = 0
    ending_inventory_units = 0
    ending_inventory_detail = {}
    # Sum stock across all locations (WH + stores)
    final_stock = {}
    for entry in D["daily_stock_log"]:
        loc_id = entry.get("location_id", entry.get("warehouse_id", ""))
        pid = entry["product_id"]
        # Keep only the latest entry per (loc, product) — last date = final stock
        final_stock[(loc_id, pid)] = entry["qty_on_hand"]
    for (loc_id, pid), qty in final_stock.items():
        if qty > 0:
            cost_val = qty * product_cost.get(pid, 0)
            ending_inventory_value += cost_val
            ending_inventory_units += qty
            if pid not in ending_inventory_detail:
                ending_inventory_detail[pid] = {"qty": 0, "value": 0}
            ending_inventory_detail[pid]["qty"] += qty
            ending_inventory_detail[pid]["value"] += cost_val

    gross_profit = round(total_rev - total_cogs)
    net_profit = round(total_net)
    ending_inv = round(ending_inventory_value)
    total_variable = round(total_var)
    # BizBot Bash score: net profit minus capital charge on ending inventory
    # Capital charge rate: 20% (represents annual cost of capital tied up in inventory)
    CAPITAL_CHARGE_RATE = 0.20
    bizbotbash_score = net_profit - round(ending_inv * CAPITAL_CHARGE_RATE)

    # ── Variable cost breakdown by type ──
    total_po_var = 0
    total_transfer_var = 0
    for po in D["po_log"]:
        total_po_var += po.get("variable_cost", 0)
    for tr in D["transfer_log"]:
        total_transfer_var += tr.get("variable_cost", 0)

    compact["ending_inventory"] = {
        "total_units": ending_inventory_units,
        "total_value": ending_inv,
        "by_product": {pid: {"qty": v["qty"], "value": round(v["value"])}
                       for pid, v in sorted(ending_inventory_detail.items(),
                                            key=lambda x: x[1]["value"], reverse=True)[:20]}
    }

    # Count unique POs and shipments (unique IDs) vs total line items
    unique_pos = len(set(po["po_id"] for po in D["po_log"])) if D["po_log"] else 0
    unique_shipments = len(set(t["transfer_id"] for t in D["transfer_log"])) if D["transfer_log"] else 0

    compact["summary"] = {
        "total_units_sold": total_units,
        "total_revenue": round(total_rev),
        "total_cogs": round(total_cogs),
        "total_fixed_costs": round(total_fixed),
        "total_variable_costs": total_variable,
        "total_po_variable_costs": round(total_po_var),
        "total_transfer_variable_costs": round(total_transfer_var),
        "total_gross_profit": gross_profit,
        "total_net_profit": net_profit,
        "total_po_lines": len(D["po_log"]),
        "total_pos": unique_pos,
        "total_transfer_lines": len(D["transfer_log"]),
        "total_transfers": unique_shipments,
        "ending_inventory_value": ending_inv,
        "ending_inventory_units": ending_inventory_units,
        "bizbotbash_score": bizbotbash_score,
    }

    return compact


# ═══════════════════════════════════════════════════════════
# SIMULATION ENGINE v2
# ═══════════════════════════════════════════════════════════
class SimulationEngine:
    def __init__(self, cfg, mode="auto"):
        """mode: 'auto' (fully autonomous) or 'bot' (step-by-step with commands)"""
        self.cfg = cfg
        self.mode = mode
        company = cfg["company"]
        self.current_date = date.fromisoformat(company["sim_start"])
        self.sim_months = company["sim_months"]
        self.working_days_set = set(company["working_days"])
        # End on the last calendar day of the final target month so the last
        # bucket in the Monthly Revenue chart isn't truncated to ~20 days.
        _start = date.fromisoformat(company["sim_start"])
        _final_year = _start.year + (_start.month - 1 + self.sim_months) // 12
        _next_month = ((_start.month - 1 + self.sim_months) % 12) + 1
        self.end_date = date(_final_year, _next_month, 1) - timedelta(days=1)

        # Separate RNG streams so bot-mode commands don't shift the sales sequence
        seed = company.get("random_seed", 2026)
        self.rng_sales = random.Random(seed)       # customer traffic, product picks, qty
        self.rng_logistics = random.Random(seed + 1)  # PO lead times, transfer transit, reliability
        self.rng_trends = random.Random(seed + 2)  # random trend events (hidden from bots)
        random.seed(seed)  # keep global seed for backward compat (load_config uses it)

        # Hidden trend-event state (bots cannot observe directly)
        self.trend_events = []        # currently active events
        self.trend_events_log = []    # every event ever rolled, for post-mortem export
        self._last_trend_month = -1

        products = cfg["products"]
        self.product_map = {p["id"]: p for p in products}
        self.product_volume = {p["id"]: p.get("volume_cm3", 5000) for p in products}
        self.refill_num = {p["id"]: p.get("refill_num", 5) for p in products}
        self.supplier_map = {s["id"]: s for s in cfg["suppliers"]}

        self.store_capacity_cm3 = {}
        for loc in cfg["physical_locs"]:
            self.store_capacity_cm3[loc["id"]] = loc.get("effective_capacity_cm3", 5_000_000)

        # Product footprint (cm²) — derived from dimensions if the explicit
        # base_area_cm2 field is absent, min 20 cm² so tiny items don't
        # dominate shelf capacity.
        self.product_area = {
            p["id"]: max(20, p.get("base_area_cm2",
                                   p.get("dim_l_cm", 10) * p.get("dim_w_cm", 10)))
            for p in products
        }

        # Per-shelf area (cm²) at each store: slots_per_shelf × SLOT_AREA_CM2.
        self.shelf_area_cm2 = {
            loc["id"]: loc.get("slots_per_shelf", 30) * SLOT_AREA_CM2
            for loc in cfg["physical_locs"]
        }
        # Number of shelves per grade at each store.
        self.grade_shelf_count = {}
        for loc in cfg["physical_locs"]:
            grades = loc.get("shelf_grades", ["B"])
            self.grade_shelf_count[loc["id"]] = {
                "A": grades.count("A"),
                "B": grades.count("B"),
                "C": grades.count("C"),
            }

        self.shelf_mult = cfg["shelf_mult"]
        self.location_mult = cfg["location_mult"]
        self.product_mult = cfg["product_mult"]
        self.location_grade = cfg["location_grade"]
        self.product_grade = cfg["product_grade"]
        self.product_shelf_grade = dict(cfg["product_shelf_grade"])  # mutable copy

        # ── Discounts: (product_id, location_id) → discount_pct ──
        # location_id = "" means global discount for that product
        self.discounts = {}

        # Initialize warehouse stock
        self.stock = {}           # total at each (location, product)
        self.shelf_stock = {}     # sellable subset at (store, product). Backroom = stock - shelf.
        for p in products:
            inv = cfg["inventory_params"][p["id"]]
            self.stock[("WH-01", p["id"])] = inv["initial_wh01"]

        # Initialize store stock
        for loc in cfg["physical_locs"]:
            loc_id = loc["id"]
            capacity_cm3 = self.store_capacity_cm3[loc_id]
            used_cm3 = 0
            shuffled = list(products)
            self.rng_logistics.shuffle(shuffled)
            for p in shuffled:
                refill = self.refill_num[p["id"]]
                vol_per_unit = self.product_volume[p["id"]]
                space_left_cm3 = capacity_cm3 - used_cm3
                max_by_volume = int(space_left_cm3 / vol_per_unit) if vol_per_unit > 0 else 0
                can_place = min(refill, max_by_volume)
                if can_place <= 0:
                    self.stock[(loc_id, p["id"])] = 0
                    self.shelf_stock[(loc_id, p["id"])] = 0
                    continue
                wh_avail = self.stock.get(("WH-01", p["id"]), 0)
                actual = min(can_place, wh_avail)
                self.stock[(loc_id, p["id"])] = actual
                # Fill the shelf first; anything beyond shelf capacity sits in backroom.
                self.shelf_stock[(loc_id, p["id"])] = min(
                    actual, self.shelf_cap_for(p["id"], loc_id)
                )
                self.stock[("WH-01", p["id"])] = wh_avail - actual
                used_cm3 += actual * vol_per_unit

        self.pending_pos = []
        self.po_counter = 0
        self.pending_transfers = []
        self.transfer_counter = 0
        self.daily_stock_log = []
        self.order_log = []
        self.po_log = []
        self.transfer_log = []
        self.financial_log = []
        self.event_log = []
        self.action_log = []  # tracks all bot + auto actions
        self.order_counter = 0
        self.day_count = 0
        self.total_revenue = 0
        self.total_cogs = 0
        self.total_variable_costs = 0   # PO + transfer variable costs
        self.daily_variable_costs = 0   # reset each day in step_day
        self._last_month = None

        # Variable cost parameters
        vc = cfg.get("variable_costs", {})
        self.po_processing_fee = vc.get("po_processing_fee", 0)
        self.po_freight_pct = vc.get("po_freight_pct", 0)
        self.transfer_cost_local = vc.get("transfer_cost_local", 0)
        self.transfer_cost_upcountry = vc.get("transfer_cost_upcountry", 0)
        self.transfer_cost_per_unit = vc.get("transfer_cost_per_unit", 0)

    # ── Helpers ──
    @staticmethod
    def _date_lte(a, b):
        if isinstance(a, str): a = date.fromisoformat(a)
        if isinstance(b, str): b = date.fromisoformat(b)
        return a <= b

    def month_index(self):
        d = self.current_date
        s = date.fromisoformat(self.cfg["company"]["sim_start"])
        return (d.year - s.year) * 12 + (d.month - s.month)

    def calendar_month(self):
        return self.current_date.month

    def is_working_day(self):
        return self.current_date.weekday() in self.working_days_set

    def get_discount(self, product_id, location_id=""):
        """Get effective discount for a product at a location.
        Checks location-specific first, then global, then 0."""
        d = self.discounts.get((product_id, location_id), None)
        if d is not None:
            return d
        d = self.discounts.get((product_id, ""), None)
        if d is not None:
            return d
        return 0.0

    def shelf_cap_for(self, product_id, location_id):
        """Max units of product P that fit on shelf at store L, given
        P's current grade assignment at that store.
            units_per_shelf = floor(shelf_area / product_base_area)
            total_cap       = units_per_shelf × shelves_at_that_grade
        Returns 0 if P isn't assigned at L, or the store has no shelves
        of the needed grade (e.g. C-grade product at a store with no
        C-shelves)."""
        if location_id not in self.shelf_area_cm2:
            return 0
        grade = self.product_shelf_grade.get((location_id, product_id), "B")
        num_shelves = self.grade_shelf_count.get(location_id, {}).get(grade, 0)
        if num_shelves == 0:
            return 0
        area_per_unit = self.product_area.get(product_id, 1)
        per_shelf = int(self.shelf_area_cm2[location_id] // area_per_unit)
        return per_shelf * num_shelves

    # ── Random trend events (hidden from bots) ────────────────
    def _targets_for_trend_scope(self, scope):
        if scope in ("location",):
            return [l["id"] for l in self.cfg["sales_locations"]]
        if scope in ("product_global", "product_local"):
            return [p["id"] for p in self.cfg["products"]]
        if scope in ("brand_global", "brand_local"):
            # TEMPORARY: one supplier == one brand. In reality suppliers
            # can carry multiple brands. Revisit when products grow a
            # `brand` field.
            return [s["id"] for s in self.cfg["suppliers"]]
        if scope in ("category_global", "category_local"):
            cats = self.cfg["categories"]
            return list(cats.keys()) if isinstance(cats, dict) else [c["id"] for c in cats]
        return []

    def _all_location_ids(self):
        return [l["id"] for l in self.cfg["sales_locations"]]

    def _product_supplier_id(self, pid):
        return self.cfg.get("product_supplier", {}).get(pid)

    def _product_category(self, pid):
        p = self.product_map.get(pid, {})
        return p.get("cat")

    def _emit_trend_event(self, mi, scope, target_id, location_id):
        """Roll permanence/duration/direction/magnitude and append the event."""
        permanent = self.rng_trends.random() < TREND_PERMANENT_PROB
        if permanent:
            duration = None
            end_month = TREND_PERMANENT_SENTINEL
            bucket = "permanent"
        else:
            r = self.rng_trends.random()
            cum = 0
            dmin, dmax = 1, 3
            for prob, lo, hi in TREND_DURATION_BUCKETS:
                cum += prob
                if r < cum:
                    dmin, dmax = lo, hi
                    break
            duration = self.rng_trends.randint(dmin, dmax)
            end_month = mi + duration
            bucket = (
                "short" if duration <= 3 else
                "mid" if duration <= 6 else "long"
            )
        direction = self.rng_trends.choice([-1, +1])
        magnitude = self.rng_trends.uniform(*TREND_MAG_RANGE)
        event = {
            "scope": scope,
            "locality": "local" if location_id is not None else "global",
            "target_id": target_id,
            "location_id": location_id,  # None for global scope events
            "direction": direction,
            "magnitude": round(magnitude, 3),
            "start_month": mi,
            "end_month": end_month,
            "duration_months": duration,
            "bucket": bucket,
            "permanent": permanent,
            "started_on": str(self.current_date),
        }
        self.trend_events.append(event)
        self.trend_events_log.append(event)

    def _roll_new_trends(self, mi):
        """At the start of each new month, roll dice for new trend events
        across every scope (global and local)."""
        all_locs = self._all_location_ids()
        for scope, per_location, rate in TREND_SCOPES:
            targets = self._targets_for_trend_scope(scope)
            locs = all_locs if per_location else [None]
            for target_id in targets:
                for loc_id in locs:
                    if self.rng_trends.random() >= rate:
                        continue
                    # For "location" scope, the target IS a location;
                    # emit with location_id=None (locality is implicit).
                    self._emit_trend_event(mi, scope, target_id, loc_id)

    def _prune_expired_trends(self, mi):
        self.trend_events = [e for e in self.trend_events if e["end_month"] >= mi]

    def _trend_multiplier(self, pid, loc_id, mi):
        """Combined trend-event multiplier for a (product, location) pairing.
        ADDITIVE combining: sum every matching event's delta, then apply
        once. Global and local events at the same scope just add. Capped
        at [0.25, 4.0] so stacked shocks can't run away."""
        if not self.trend_events:
            return 1.0
        sup_id = self._product_supplier_id(pid)
        cat_id = self._product_category(pid)
        total_delta = 0.0
        for e in self.trend_events:
            if not (e["start_month"] <= mi <= e["end_month"]):
                continue
            scope = e["scope"]
            tgt = e["target_id"]
            event_loc = e.get("location_id")

            # Match logic — each scope has a target and, for *_local,
            # must also match the location.
            if scope == "location":
                if tgt != loc_id:
                    continue
            elif scope == "product_global":
                if tgt != pid:
                    continue
            elif scope == "product_local":
                if tgt != pid or event_loc != loc_id:
                    continue
            elif scope == "brand_global":
                if tgt != sup_id:
                    continue
            elif scope == "brand_local":
                if tgt != sup_id or event_loc != loc_id:
                    continue
            elif scope == "category_global":
                if tgt != cat_id:
                    continue
            elif scope == "category_local":
                if tgt != cat_id or event_loc != loc_id:
                    continue
            else:
                continue

            total_delta += e["direction"] * e["magnitude"]

        mult = 1.0 + total_delta
        lo, hi = TREND_MULT_CAP
        return max(lo, min(hi, mult))

    def get_demand(self, product_id, location_id=None, include_discount=True):
        """Calculate daily demand with optional price elasticity from discounts."""
        hv = self.cfg["hidden_vars"][product_id]
        mi = self.month_index()
        cm = self.calendar_month() - 1

        trend_mult = _popularity_cycle(product_id, mi, hv)
        base = hv["base_daily_demand"] * trend_mult
        base *= self._trend_multiplier(product_id, location_id or "", mi)
        season_mult = hv["seasonality_12m"][cm]

        hype_mult = 1.0
        for (hm, hm_mult) in hv["hype_events"]:
            if mi == hm:
                hype_mult = hm_mult
            elif mi == hm + 1:
                hype_mult = max(hype_mult, 1.0 + (hm_mult - 1.0) * 0.3)

        buzz_bonus = 1.0 + (hv["social_media_buzz"] - 0.5) * 0.3 * (hype_mult - 1.0)
        comp_penalty = 1.0 - hv["competitor_pressure"] * 0.15
        dow = self.current_date.weekday()
        dow_mult = {0: 0.85, 1: 0.90, 2: 1.0, 3: 1.05, 4: 1.10, 5: 1.20}.get(dow, 1.0)

        if base > hv["market_saturation_threshold"] / 30:
            base = base * 0.92

        demand = base * season_mult * hype_mult * buzz_bonus * comp_penalty * dow_mult

        # ABC scoring (physical stores)
        if location_id and location_id in self.location_grade:
            loc_grade = self.location_grade[location_id]
            shelf_grade = self.product_shelf_grade.get((location_id, product_id), "B")
            prod_grade = self.product_grade.get(product_id, "B")
            demand *= self.location_mult.get(loc_grade, 1.0)
            demand *= self.shelf_mult.get(shelf_grade, 1.0)
            demand *= self.product_mult.get(prod_grade, 1.0)

        # Price elasticity from discount
        if include_discount:
            disc = self.get_discount(product_id, location_id or "")
            if disc > 0:
                elasticity = hv.get("price_sensitivity", 0.3) * 3.0
                demand *= (1 - disc) ** (-elasticity)

        noise = self.rng_sales.gauss(1.0, 0.15)
        demand = max(0, demand * noise)
        return int(round(demand))

    def _store_used_volume(self, loc_id):
        total = 0
        for p in self.cfg["products"]:
            qty = self.stock.get((loc_id, p["id"]), 0)
            total += qty * self.product_volume[p["id"]]
        return total

    def _store_free_volume(self, loc_id):
        return self.store_capacity_cm3[loc_id] - self._store_used_volume(loc_id)

    def _max_units_that_fit(self, loc_id, product_id):
        free = self._store_free_volume(loc_id)
        vol = self.product_volume[product_id]
        return int(free / vol) if vol > 0 else 0

    # ── Auto mode: generate commands identical to demo_baseline_bot ──
    def _generate_auto_commands(self):
        """Build commands matching the demo baseline bot logic exactly."""
        commands = []
        products = self.cfg["products"]
        inv_params = self.cfg["inventory_params"]
        product_supplier = self.cfg["product_supplier"]

        # ── 1. WH-01 reorder: order-up-to target when stock drops below trigger ──
        # Trigger = 1/3 of target (per product, not a flat "20 units" threshold).
        # Qty = target - on_hand so POs taper as WH refills.
        # Group by supplier, skip batch if total < ฿20,000 THB.
        pending_product_ids = set(
            po["product_id"] for po in self.pending_pos if not po["received"])

        po_by_supplier = defaultdict(list)  # supplier_id → [(product, qty, cost)]
        for p in products:
            pid = p["id"]
            if pid in pending_product_ids:
                continue
            params = inv_params[pid]
            target = params.get("target_wh", params.get("reorder_qty", 40))
            trigger = params.get("trigger_wh", max(2, target // 3))
            on_hand = self.stock.get(("WH-01", pid), 0)
            if on_hand >= trigger:
                continue
            sup_id = product_supplier.get(pid)
            if not sup_id:
                continue
            qty = max(1, target - on_hand)
            cost = qty * p["cost"]
            po_by_supplier[sup_id].append((pid, qty, cost))

        for sup_id, items in po_by_supplier.items():
            total_cost = sum(cost for _, _, cost in items)
            if total_cost < 20000:
                continue
            po_items = [{"product_id": pid, "qty": qty} for pid, qty, _ in items]
            commands.append({
                "action": "issue_po",
                "supplier_id": sup_id,
                "items": po_items,
            })

        # ── 2. Store refill: transfer from WH when store stock = 0 ──
        # Group all products going to the same store into one shipment
        pending_transfer_keys = set(
            (t["product_id"], t["to_loc"]) for t in self.pending_transfers if not t["received"])

        # Track WH stock locally to avoid double-transferring
        wh_stock_local = {}
        for p in products:
            wh_stock_local[p["id"]] = self.stock.get(("WH-01", p["id"]), 0)

        # Collect items per destination store
        store_items = defaultdict(list)  # loc_id → [{product_id, qty}, ...]
        for loc in self.cfg["physical_locs"]:
            loc_id = loc["id"]
            for p in products:
                pid = p["id"]
                store_qty = self.stock.get((loc_id, pid), 0)
                if store_qty > 0:
                    continue
                if (pid, loc_id) in pending_transfer_keys:
                    continue
                refill = self.refill_num[pid]
                wh_avail = wh_stock_local.get(pid, 0)
                actual = min(refill, wh_avail)
                if actual > 0:
                    store_items[loc_id].append({"product_id": pid, "qty": actual})
                    wh_stock_local[pid] = wh_avail - actual

        # Emit one grouped transfer command per store
        for loc_id, items in store_items.items():
            commands.append({
                "action": "transfer",
                "from_loc": "WH-01",
                "to_loc": loc_id,
                "items": items,
            })

        return commands

    # ── Command execution (shared by auto + bot modes) ──
    def _execute_commands(self, commands):
        """Process a list of command dicts. Used by both auto and bot modes.

        Transfer commands support two formats:
          1. Grouped (recommended): one truck, multiple products, one base fee
             {"action": "transfer", "from_loc": "WH-01", "to_loc": "LOC-03",
              "items": [{"product_id": "PRD-001", "qty": 5}, {"product_id": "PRD-005", "qty": 3}]}
          2. Single (legacy): one product per command, each gets its own transfer ID & base fee
             {"action": "transfer", "from_loc": "WH-01", "to_loc": "LOC-03",
              "product_id": "PRD-001", "qty": 5}

        Bots that group transfers into fewer shipments pay fewer base fees.
        """
        source = self.mode  # "auto" or "bot"

        for cmd in commands:
            action = cmd.get("action", "")

            if action == "issue_po":
                sup_id = cmd["supplier_id"]
                sup = self.supplier_map.get(sup_id)
                if not sup:
                    continue

                # Support both grouped and single format
                if "items" in cmd:
                    raw_items = cmd["items"]  # [{product_id, qty}, ...]
                else:
                    raw_items = [{"product_id": cmd["product_id"], "qty": cmd["qty"]}]

                # Validate items
                valid_items = []
                for item in raw_items:
                    pid = item["product_id"]
                    qty = item["qty"]
                    if qty > 0 and pid in self.product_map:
                        valid_items.append((pid, qty))

                if not valid_items:
                    continue

                # One PO ID per command = one purchase order, one processing fee
                lead = self.rng_logistics.randint(sup["lead_days"][0], sup["lead_days"][1])
                arrival = self.current_date + timedelta(days=lead)
                self.po_counter += 1
                po_id = f"PO-{self.po_counter:06d}"

                # Calculate total PO value for freight
                total_po_value = sum(qty * self.product_map[pid]["cost"] for pid, qty in valid_items)

                # Variable costs: one processing fee per PO + freight on total value
                po_var_cost = self.po_processing_fee + (total_po_value * self.po_freight_pct)
                self.daily_variable_costs += po_var_cost

                # Create one PO record per product line, all sharing the same po_id
                # Distribute variable cost: processing fee on first item, freight on each
                product_details = []
                for i, (pid, qty) in enumerate(valid_items):
                    p = self.product_map[pid]
                    line_cost = qty * p["cost"]
                    if i == 0:
                        line_var_cost = self.po_processing_fee + (line_cost * self.po_freight_pct)
                    else:
                        line_var_cost = line_cost * self.po_freight_pct
                    po = {
                        "po_id": po_id, "date": self.current_date,
                        "supplier_id": sup_id, "product_id": pid,
                        "qty_ordered": qty, "qty_received": 0,
                        "unit_cost": p["cost"], "total_cost": line_cost,
                        "arrival_date": arrival, "lead_days": lead,
                        "received": False, "status": "pending", "source": source,
                        "variable_cost": round(line_var_cost, 2),
                    }
                    self.pending_pos.append(po)
                    self.po_log.append(po)
                    product_details.append(f"{qty}× {pid}")

                self.action_log.append({"date": self.current_date, "source": source,
                    "action": "issue_po",
                    "details": f"PO {po_id}: {', '.join(product_details)} from {sup_id}, "
                               f"ETA {lead}d, {len(valid_items)} products, "
                               f"total={total_po_value:,.0f}, var_cost={po_var_cost:.0f}"})

            elif action == "transfer":
                from_loc = cmd["from_loc"]
                to_loc = cmd["to_loc"]

                # Build items list — supports both grouped and single format
                if "items" in cmd:
                    raw_items = cmd["items"]  # [{product_id, qty}, ...]
                else:
                    raw_items = [{"product_id": cmd["product_id"], "qty": cmd["qty"]}]

                # Validate stock and deduct for each item
                valid_items = []
                for item in raw_items:
                    pid = item["product_id"]
                    qty = item["qty"]
                    avail = self.stock.get((from_loc, pid), 0)
                    actual = min(qty, avail)
                    if actual <= 0:
                        continue
                    self.stock[(from_loc, pid)] = avail - actual
                    valid_items.append((pid, actual))

                if not valid_items:
                    continue

                # One transfer ID per command = one truck
                is_upcountry = any(loc.get("region", "").startswith("Upcountry")
                                   for loc in self.cfg["sales_locations"] if loc["id"] == to_loc)
                transit = self.rng_logistics.randint(2, 3) if is_upcountry else 1
                transfer_type = "WH→Store" if from_loc == "WH-01" else "Store→Store"
                self.transfer_counter += 1
                trf_id = f"TRF-{self.transfer_counter:05d}"

                # Variable costs: one base fee per shipment + per-unit handling for all units
                base_transfer_fee = self.transfer_cost_upcountry if is_upcountry else self.transfer_cost_local
                total_units = sum(qty for _, qty in valid_items)
                shipment_var_cost = base_transfer_fee + (total_units * self.transfer_cost_per_unit)
                self.daily_variable_costs += shipment_var_cost

                # Create one transfer record per product line, all sharing the same transfer_id
                # Distribute variable cost: base fee on first item, handling on each
                product_details = []
                for i, (pid, actual) in enumerate(valid_items):
                    if i == 0:
                        line_var_cost = base_transfer_fee + (actual * self.transfer_cost_per_unit)
                    else:
                        line_var_cost = actual * self.transfer_cost_per_unit
                    t = {
                        "transfer_id": trf_id,
                        "date": self.current_date, "product_id": pid,
                        "from_loc": from_loc, "to_loc": to_loc,
                        "transfer_type": transfer_type,
                        "qty": actual,
                        "arrival_date": self.current_date + timedelta(days=transit),
                        "received": False, "source": source,
                        "variable_cost": round(line_var_cost, 2),
                    }
                    self.pending_transfers.append(t)
                    self.transfer_log.append(t)
                    product_details.append(f"{actual}× {pid}")

                self.action_log.append({"date": self.current_date, "source": source,
                    "action": "transfer",
                    "details": f"TRF {trf_id}: {', '.join(product_details)} from {from_loc} → {to_loc}, "
                               f"{len(valid_items)} products, {total_units} units, var_cost={shipment_var_cost:.0f}"})

            elif action == "set_discount":
                pid = cmd["product_id"]
                disc = cmd["discount_pct"]
                loc_id = cmd.get("location_id", "")
                self.discounts[(pid, loc_id)] = disc
                self.action_log.append({"date": self.current_date, "source": source,
                    "action": "set_discount",
                    "details": f"{pid} @ {loc_id or 'ALL'}: {disc*100:.1f}% off"})

            elif action == "set_shelf":
                loc_id = cmd["location_id"]
                pid = cmd["product_id"]
                grade = cmd["shelf_grade"]
                self.product_shelf_grade[(loc_id, pid)] = grade
                # If the new grade has a smaller cap (e.g. demotion A → C
                # where the store has fewer C-shelves), trim visible stock
                # to the new cap — the excess implicitly moves to backroom.
                new_cap = self.shelf_cap_for(pid, loc_id)
                if self.shelf_stock.get((loc_id, pid), 0) > new_cap:
                    self.shelf_stock[(loc_id, pid)] = new_cap
                self.action_log.append({"date": self.current_date, "source": "bot",
                    "action": "set_shelf",
                    "details": f"{pid} @ {loc_id} → shelf grade {grade}"})

    # ── Single day step ──
    def step_day(self, commands=None):
        """Advance simulation by one day. In bot mode, pass commands to execute.
        Returns True if simulation should continue, False if done."""
        mi = self.month_index()
        if mi >= self.sim_months or self.current_date > self.end_date:
            return False

        products = self.cfg["products"]
        suppliers = self.cfg["suppliers"]
        warehouses = self.cfg["warehouses"]
        sales_locations = {loc["id"]: loc for loc in self.cfg["sales_locations"]}
        physical_loc_ids = set(l["id"] for l in self.cfg["physical_locs"])
        online_loc_ids = set(l["id"] for l in self.cfg["online_locs"])
        inv_params = self.cfg["inventory_params"]
        product_supplier = self.cfg["product_supplier"]
        monthly_costs = self.cfg["monthly_costs"]
        total_monthly_fixed = sum(monthly_costs.values())

        # Reset daily variable costs (accumulated by _execute_commands)
        self.daily_variable_costs = 0

        # Check for month boundary → write Excel
        current_month_str = self.current_date.strftime("%Y-%m")
        if self._last_month and self._last_month != current_month_str:
            write_monthly_excel(self, self._last_month)
        self._last_month = current_month_str

        # First day of a new sim-month: roll new random trend events,
        # expire old ones. Deterministic via self.rng_trends (seed + 2).
        if mi != self._last_trend_month:
            self._roll_new_trends(mi)
            self._prune_expired_trends(mi)
            self._last_trend_month = mi

        # In auto mode, generate commands just like a bot would (before sales)
        if self.mode == "auto":
            auto_commands = self._generate_auto_commands()
            self._execute_commands(auto_commands)
        elif commands:
            self._execute_commands(commands)

        if not self.is_working_day():
            self.current_date += timedelta(days=1)
            return True

        self.day_count += 1
        day_revenue = 0
        day_cogs = 0
        day_units_sold = 0
        day_orders = 0
        day_stockouts = 0

        # ── 1. Receive PO deliveries ──
        for po in self.pending_pos:
            if self._date_lte(po["arrival_date"], self.current_date) and not po["received"]:
                po["received"] = True
                sup = self.supplier_map.get(po["supplier_id"])
                if sup and self.rng_logistics.random() > sup["reliability"]:
                    po["qty_received"] = int(po["qty_ordered"] * self.rng_logistics.uniform(0.6, 0.9))
                    po["status"] = "partial"
                else:
                    po["qty_received"] = po["qty_ordered"]
                    po["status"] = "complete"
                self.stock[("WH-01", po["product_id"])] = (
                    self.stock.get(("WH-01", po["product_id"]), 0) + po["qty_received"])
                self.event_log.append({"date": self.current_date, "type": "delivery",
                    "ref": po["po_id"],
                    "detail": f"Received {po['qty_received']}/{po['qty_ordered']} of {po['product_id']}"})

        # ── 2. Receive pending transfers ──
        for t in self.pending_transfers:
            if self._date_lte(t["arrival_date"], self.current_date) and not t["received"]:
                t["received"] = True
                dest = t["to_loc"]
                self.stock[(dest, t["product_id"])] = (
                    self.stock.get((dest, t["product_id"]), 0) + t["qty"])

        # ── 3. Retail sales ──
        for loc_id, loc_info in sales_locations.items():
            is_physical = loc_id in physical_loc_ids
            is_online = loc_id in online_loc_ids

            working_days = loc_info.get("working_days", None)
            if working_days is not None:
                if self.current_date.weekday() not in working_days:
                    continue

            traffic = loc_info.get("daily_foot_traffic", 0)
            base_conv = loc_info.get("conversion_rate", 0)
            loc_g = self.location_grade.get(loc_id, "B")
            loc_mult = self.location_mult.get(loc_g, 1.0)
            adjusted_conv = base_conv * loc_mult

            dow_noise = self.rng_sales.gauss(1.0, 0.1)
            num_customers = max(0, int(traffic * adjusted_conv * dow_noise))
            if num_customers == 0:
                continue

            cm = self.calendar_month() - 1
            mi_val = self.month_index()

            demand_weights = []
            for p in products:
                hv = self.cfg["hidden_vars"][p["id"]]
                base_d = hv["base_daily_demand"]
                buzz_factor = 1.0 + hv["social_media_buzz"] * 0.5
                loyalty = hv["brand_loyalty"]
                season = hv["seasonality_12m"][cm]
                trend = _popularity_cycle(p["id"], mi_val, hv)
                event_trend = self._trend_multiplier(p["id"], loc_id, mi_val)

                w = base_d * buzz_factor * loyalty * season * trend * event_trend

                # Shelf grade affects product visibility/attractiveness at physical stores
                if is_physical:
                    shelf_grade = self.product_shelf_grade.get((loc_id, p["id"]), "B")
                    w *= self.shelf_mult.get(shelf_grade, 1.0)

                # Discount boosts product weight (makes discounted items more attractive)
                disc = self.get_discount(p["id"], loc_id)
                if disc > 0:
                    elasticity = hv.get("price_sensitivity", 0.3) * 3.0
                    w *= (1 - disc) ** (-elasticity)

                demand_weights.append(max(w, 0.1))

            for _ in range(num_customers):
                self.order_counter += 1
                order_id = f"ORD-{self.order_counter:06d}"
                cust_name = f"Customer {self.order_counter}"

                n_items = self.rng_sales.choices([1, 2, 3], weights=[50, 35, 15])[0]
                selected = self.rng_sales.choices(products, weights=demand_weights, k=n_items)
                seen = set()
                unique_selected = []
                for p in selected:
                    if p["id"] not in seen:
                        seen.add(p["id"])
                        unique_selected.append(p)

                for prod in unique_selected:
                    qty_wanted = self.rng_sales.choices([1, 2], weights=[70, 30])[0]
                    disc = self.get_discount(prod["id"], loc_id)
                    sell_price = round(prod["price"] * (1 - disc), 2)

                    if is_physical:
                        # Customers can only buy what's ON THE SHELF. If the
                        # shelf is empty the sale is lost even when backroom
                        # has stock — bots must replenish shelf faster.
                        shelf_avail = self.shelf_stock.get((loc_id, prod["id"]), 0)
                        qty_filled = min(qty_wanted, shelf_avail)
                        self.shelf_stock[(loc_id, prod["id"])] = shelf_avail - qty_filled
                        self.stock[(loc_id, prod["id"])] = self.stock.get((loc_id, prod["id"]), 0) - qty_filled
                    else:
                        wh_avail = self.stock.get(("WH-01", prod["id"]), 0)
                        qty_filled = min(qty_wanted, wh_avail)
                        self.stock[("WH-01", prod["id"])] = wh_avail - qty_filled

                    if qty_filled > 0:
                        line_total = round(qty_filled * sell_price, 2)
                        day_revenue += line_total
                        day_cogs += qty_filled * prod["cost"]
                        day_units_sold += qty_filled

                    qty_backordered = qty_wanted - qty_filled
                    if qty_backordered > 0:
                        day_stockouts += 1

                    self.order_log.append({
                        "order_id": order_id, "date": self.current_date,
                        "customer_id": cust_name, "product_id": prod["id"],
                        "qty_ordered": qty_wanted, "qty_filled": qty_filled,
                        "qty_backordered": qty_backordered,
                        "unit_price": sell_price, "discount_pct": disc,
                        "line_total": round(qty_filled * sell_price, 2),
                        "cogs": round(qty_filled * prod["cost"], 2),
                        "fulfill_warehouse": "STORE" if is_physical else "WH-01",
                        "sales_location_id": loc_id,
                        "sales_location_name": loc_info.get("name", loc_id),
                        "sales_region": loc_info.get("region", "Unknown"),
                        "status": "filled" if qty_backordered == 0 else (
                            "partial" if qty_filled > 0 else "backordered"),
                        "source": "auto",
                    })

                day_orders += 1

        self.total_revenue += day_revenue
        self.total_cogs += day_cogs
        self.total_variable_costs += self.daily_variable_costs

        # ── 4a. End-of-day shelf refill from backroom ──
        # After sales, any product whose shelf is below its per-SKU cap
        # gets restocked from the store's backroom. Backroom is the
        # implicit gap between self.stock and self.shelf_stock.
        for loc in self.cfg["physical_locs"]:
            loc_id = loc["id"]
            for p in products:
                pid = p["id"]
                total = self.stock.get((loc_id, pid), 0)
                on_shelf = self.shelf_stock.get((loc_id, pid), 0)
                backroom = total - on_shelf
                if backroom <= 0:
                    continue
                cap = self.shelf_cap_for(pid, loc_id)
                need = cap - on_shelf
                if need <= 0:
                    continue
                moved = min(need, backroom)
                self.shelf_stock[(loc_id, pid)] = on_shelf + moved
                # self.stock is unchanged (total hasn't moved, just reallocated)

        # ── 4b. Stock snapshots (month-end only, post-refill view) ──
        # Per-day snapshots for 60+ months × 9 locations × 80 SKUs would
        # blow past Mongo's 16MB BSON limit. Monthly is enough detail
        # for the dashboard's stock view and the exported spreadsheet.
        next_day = self.current_date + timedelta(days=1)
        is_month_end = (next_day.month != self.current_date.month)
        is_sim_end = (next_day > self.end_date)
        if is_month_end or is_sim_end:
            for p in products:
                for wh in warehouses:
                    qty = self.stock.get((wh["id"], p["id"]), 0)
                    self.daily_stock_log.append({
                        "date": self.current_date, "location_id": wh["id"],
                        "location_type": "warehouse",
                        "product_id": p["id"], "qty_on_hand": qty,
                        "shelf_qty": 0, "backroom_qty": qty,
                    })
                for loc in self.cfg["physical_locs"]:
                    qty = self.stock.get((loc["id"], p["id"]), 0)
                    shelf = self.shelf_stock.get((loc["id"], p["id"]), 0)
                    self.daily_stock_log.append({
                        "date": self.current_date, "location_id": loc["id"],
                        "location_type": "store",
                        "product_id": p["id"], "qty_on_hand": qty,
                        "shelf_qty": shelf, "backroom_qty": qty - shelf,
                    })

        # ── 7. Daily financials ──
        daily_fixed = total_monthly_fixed / 26
        day_var_costs = self.daily_variable_costs
        day_total_costs = day_cogs + daily_fixed + day_var_costs
        self.financial_log.append({
            "date": self.current_date, "month": self.current_date.strftime("%Y-%m"),
            "revenue": round(day_revenue, 2), "cogs": round(day_cogs, 2),
            "gross_profit": round(day_revenue - day_cogs, 2),
            "fixed_costs": round(daily_fixed, 2),
            "variable_costs": round(day_var_costs, 2),
            "net_profit": round(day_revenue - day_total_costs, 2),
            "orders_count": day_orders, "units_sold": day_units_sold,
            "stockout_events": day_stockouts,
        })

        self.current_date += timedelta(days=1)
        return True

    # ── Full auto run ──
    def run(self):
        """Run the entire simulation in auto mode (backward-compatible)."""
        print("Writing initial state Excel...")
        write_initial_state_excel(self)
        while self.step_day():
            pass
        # Write final month
        if self._last_month:
            write_monthly_excel(self, self._last_month)

        t_types = defaultdict(int)
        for t in self.transfer_log:
            t_types[t["transfer_type"]] += 1
        print(f"Simulation complete: {self.day_count} working days")
        print(f"Orders: {self.order_counter}, POs: {self.po_counter}, Transfers: {self.transfer_counter}")
        for tt, cnt in sorted(t_types.items()):
            print(f"  {tt}: {cnt}")

    def get_output(self):
        return {
            "company": self.cfg["company"],
            "categories": self.cfg["categories"],
            "products": self.cfg["products"],
            "hidden_vars": self.cfg["hidden_vars"],
            "suppliers": self.cfg["suppliers"],
            "customers": self.cfg["customers"],
            "warehouses": self.cfg["warehouses"],
            "sales_locations": self.cfg["sales_locations"],
            "inventory_params": self.cfg["inventory_params"],
            "monthly_costs": self.cfg["monthly_costs"],
            "variable_costs": self.cfg.get("variable_costs", {}),
            "discount_tiers": self.cfg["discount_tiers"],
            "daily_stock_log": self.daily_stock_log,
            "order_log": self.order_log,
            "po_log": self.po_log,
            "transfer_log": self.transfer_log,
            "financial_log": self.financial_log,
            "event_log": self.event_log,
            "action_log": self.action_log,
            "summary": {
                "total_days": self.day_count,
                "total_orders": self.order_counter,
                "total_pos": self.po_counter,
                "total_transfers": self.transfer_counter,
            }
        }


if __name__ == "__main__":
    from datetime import date as date_type

    class DateEncoder(json.JSONEncoder):
        def default(self, obj):
            if isinstance(obj, date_type):
                return obj.isoformat()
            return super().default(obj)

    cfg = load_config()
    engine = SimulationEngine(cfg, mode="auto")
    engine.run()
    output = engine.get_output()
    out_path = os.path.join(PROJECT_DIR, "sim_output.json")
    with open(out_path, "w") as f:
        json.dump(output, f, cls=DateEncoder, default=str)
    print(f"Data saved to {out_path}")
