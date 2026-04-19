"""Generate the downloadable Excel bundle from Mongo run_raw logs.

Matches the original filesystem format: a zip containing
  initial_state.xlsx   — catalog + starting shelf state
  month_YYYY-MM.xlsx   — one file per month of activity
  final_state.xlsx     — ending shelf state + run summary
"""
import io
import zipfile
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

import mongo_runs
from db import get_db


HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="333333")


# ─────────────────────────────────────────────────────────────
# Sheet-writing helpers
# ─────────────────────────────────────────────────────────────

def _write_sheet(ws, rows, title):
    ws.title = title
    if not rows:
        ws.append([f"(no {title.lower()} in this run)"])
        return
    columns = []
    seen = set()
    for r in rows:
        for k in r.keys():
            if k not in seen:
                seen.add(k)
                columns.append(k)
    ws.append(columns)
    for c, cell in enumerate(ws[1], start=1):
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    for r in rows:
        ws.append([_coerce(r.get(k)) for k in columns])
    ws.freeze_panes = "A2"
    for col_idx, key in enumerate(columns, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(
            12, min(30, len(str(key)) + 2)
        )


def _coerce(v):
    if isinstance(v, (dict, list)):
        import json as _json
        return _json.dumps(v, default=str)
    return v


def _wb_to_bytes(wb):
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────
# Partition logs by month (YYYY-MM)
# ─────────────────────────────────────────────────────────────

_MONTHLY_LOG_KEYS = (
    "order_log",
    "po_log",
    "transfer_log",
    "daily_stock_log",
    "financial_log",
    "action_log",
)

def _group_by_month(raw):
    months = defaultdict(lambda: {k: [] for k in _MONTHLY_LOG_KEYS})
    for key in _MONTHLY_LOG_KEYS:
        for entry in raw.get(key, []) or []:
            date_str = str(entry.get("date", ""))
            if len(date_str) >= 7:
                m = date_str[:7]
                months[m][key].append(entry)
    return months


# ─────────────────────────────────────────────────────────────
# Workbook builders
# ─────────────────────────────────────────────────────────────

def _build_initial_state(raw, summary, cfg):
    """Catalog + starting shelf state. One file, many reference sheets."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Field", "Value"])
    for c in ws[1]:
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    ws.append(["label", summary.get("label", "")])
    ws.append(["months", summary.get("months", "")])
    ws.append(["sim_start", summary.get("sim_start", "")])
    ws.append(["sim_end", summary.get("sim_end", "")])
    ws.append(["bot", summary.get("bot_slug", "")])
    ws.append(["seed", summary.get("seed", "")])
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 28

    # Catalog reference tables (from challenge_configs)
    if cfg:
        _write_sheet(wb.create_sheet("Products"),  cfg.get("products", []),        "Products")
        _write_sheet(wb.create_sheet("Locations"), cfg.get("sales_locations", []), "Locations")
        _write_sheet(wb.create_sheet("Suppliers"), cfg.get("suppliers", []),       "Suppliers")

    # Shelf starting state
    _write_sheet(wb.create_sheet("Shelf_Layout"),
                 raw.get("shelf_layout", []) or [], "Shelf_Layout")
    _write_sheet(wb.create_sheet("Shelf_Assignments_Initial"),
                 raw.get("shelf_assignments_initial", []) or [],
                 "Shelf_Assignments_Initial")
    return _wb_to_bytes(wb)


def _build_monthly(logs, month_key):
    """One month of activity — sales, POs, transfers, stock snapshot, P&L, actions."""
    wb = Workbook()
    ws = wb.active
    _write_sheet(ws, logs["order_log"], "Sales")
    _write_sheet(wb.create_sheet("Purchase_Orders"), logs["po_log"],         "Purchase_Orders")
    _write_sheet(wb.create_sheet("Transfers"),       logs["transfer_log"],   "Transfers")
    _write_sheet(wb.create_sheet("Stock_Snapshot"),  logs["daily_stock_log"],"Stock_Snapshot")
    _write_sheet(wb.create_sheet("Financials"),      logs["financial_log"],  "Financials")
    _write_sheet(wb.create_sheet("Action_Log"),      logs["action_log"],     "Action_Log")
    return _wb_to_bytes(wb)


def _build_final_state(raw, summary):
    """End-of-run: KPIs, final shelf assignment, shelf_map."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Field", "Value"])
    for c in ws[1]:
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    for k in ("label", "months", "sim_start", "sim_end", "bot_slug",
              "started_at", "finished_at"):
        if k in summary:
            ws.append([k, str(summary.get(k, ""))])
    for k, v in (summary.get("summary") or {}).items():
        ws.append([k, v])
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 22

    _write_sheet(wb.create_sheet("Shelf_Assignments_Final"),
                 raw.get("shelf_assignments_final", []) or [],
                 "Shelf_Assignments_Final")
    _write_sheet(wb.create_sheet("Shelf_Map"),
                 raw.get("shelf_map", []) or [],
                 "Shelf_Map")
    return _wb_to_bytes(wb)


# ─────────────────────────────────────────────────────────────
# Public: build the zip
# ─────────────────────────────────────────────────────────────

def run_to_zip_bytes(run_id):
    """Package initial_state.xlsx + one xlsx per month + final_state.xlsx
    into a zip. Returns (zip_bytes, filename) or (None, None) if the
    requested run doesn't have raw data in Mongo."""
    raw = mongo_runs.get_run_raw(run_id)
    summary = mongo_runs.get_run_summary(run_id)
    if not raw or not summary:
        return None, None

    # Load latest challenge config for catalog reference sheets
    cfg = None
    db = get_db()
    if db is not None:
        cfg = db.challenge_configs.find_one(
            {"challenge_id": summary.get("challenge_id")},
            sort=[("version", -1)],
        )

    months = _group_by_month(raw)

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("initial_state.xlsx", _build_initial_state(raw, summary, cfg))
        for month_key in sorted(months.keys()):
            zf.writestr(f"month_{month_key}.xlsx",
                        _build_monthly(months[month_key], month_key))
        zf.writestr("final_state.xlsx", _build_final_state(raw, summary))

    label = str(summary.get("label", "run")).replace(" ", "_")
    filename = f"toyland-{label}.zip"
    return buf.getvalue(), filename
